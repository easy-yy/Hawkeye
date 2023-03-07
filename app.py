# -*- coding: utf-8 -*-

import functools
import glob
import math
import os
import os.path as osp
import re
import shutil
import webbrowser

import imgviz
import numpy as np
from tqdm import tqdm
from qtpy import QtCore
from qtpy.QtCore import Qt
from qtpy import QtGui
from qtpy import QtWidgets

from info import __appname__
from info import PY2
from info import QT5

import utils
from config import get_config
from label_file import LabelFile
from label_file import LabelFileError
from logger import logger
from shape import Shape
from widgets import BrightnessContrastDialog
from widgets import Canvas
from widgets import LabelDialog
from widgets import LabelListWidget
from widgets import LabelListWidgetItem
from widgets import ToolBar
from widgets import UniqueLabelQListWidget
from widgets import ZoomWidget
from widgets.central_widget import CentralWidget

import requests
import cv2
import segmentation
import new_seg_zhimi
import predict
import time
import xlrd
import xlwt
import copy
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as image1
from qtpy.QtWidgets import QMessageBox
from xlutils.filter import process, XLRDReader, XLWTWriter
from threading import Thread
from widgets.feature import FeatureWidget
from collections import Counter
from widgets.login_dialog import LoginDialog
from widgets.image_ import ImageList
from widgets.orb_widget import ORBWidget
from widgets.tqdm_widget import TqdmDialog, TestThread
from skimage import io
from PIL import ImageFilter, Image
from utils import resizeImage

# FIXME
# - [medium] Set max zoom value to something big enough for FitWidth/Window

# TODO(unknown):
# - [high] Add polygon movement with arrow keys
# - [high] Deselect shape when clicking and already selected(?)
# - [low,maybe] Preview images on file dialogs.
# - Zoom is too "steppy".


LABEL_COLORMAP = imgviz.label_colormap(value=200)


class MainWindow(QtWidgets.QMainWindow):

    FIT_WINDOW, FIT_WIDTH, MANUAL_ZOOM = 0, 1, 2

    def __init__(
        self,
        config=None,
        filename=None,
        output=None,
        output_file=None,
        output_dir=None,
    ):
        if output is not None:
            logger.warning(
                "argument output is deprecated, use output_file instead"
            )
            if output_file is None:
                output_file = output

        # see config/default_config.yaml for valid configuration
        if config is None:
            config = get_config()
        self._config = config
        self.feature = FeatureWidget()
        self.isLogin = True
        # self.preImage = None

        # set default shape colors
        Shape.line_color = QtGui.QColor(*self._config["shape"]["line_color"])
        Shape.fill_color = QtGui.QColor(*self._config["shape"]["fill_color"])
        Shape.select_line_color = QtGui.QColor(
            *self._config["shape"]["select_line_color"]
        )
        Shape.select_fill_color = QtGui.QColor(
            *self._config["shape"]["select_fill_color"]
        )
        Shape.vertex_fill_color = QtGui.QColor(
            *self._config["shape"]["vertex_fill_color"]
        )
        Shape.hvertex_fill_color = QtGui.QColor(
            *self._config["shape"]["hvertex_fill_color"]
        )

        ip = self._config['server']['ip']
        port = self._config['server']['port']
        self.server_path = f'http://{ip}:{port}'

        self.pelletFeatures = {
            '解理': '无',
            '消光': '无',
            '晶形': '无',
            '颜色': '无',
            '类型': '无',
            '特征': '无',
            '突起': '无',
            '干涉色': '无',
            '风化特征': '无',
            '双晶特征': '无',
            '组成特征': '无',
            '镜下特征': '无',
            '结构特征': '无',
            '成分特征':  '无',
            '构造特征': '无',
            '其他': '无',
            # 保存颗粒ID
            'feature_id': '',
            # 保存颗粒长径
            'facePorosity': '',
            # 保存颗粒面积
            'faceArea': '',
            # 颗粒标签
            'labelName': '默认名',
        }
        self.labelsDict = (
            "_background_",
            "石英",
            "燧石",
            "微斜长石",
            "正长石",
            "条纹长石",
            "斜长石",
            "沉积岩屑",
            "石英岩岩屑",
            "变质岩屑",
            "花岗岩岩屑",
            "安山岩岩屑",
            "玄武岩岩屑",
            "酸性喷出岩岩屑",
            "火山碎屑岩岩屑",
            "云母",
            "其它"
        )

        super(MainWindow, self).__init__()
        self.setWindowTitle(__appname__)

        # Whether we need to save or not.
        self.dirty = False

        self._noSelectionSlot = False
        self.orb = ORBWidget()

        # Main widgets and related state.
        self.labelDialog = LabelDialog(
            parent=self,
            labels=self._config["labels"],
            sort_labels=self._config["sort_labels"],
            show_text_field=self._config["show_label_text_field"],
            completion=self._config["label_completion"],
            fit_to_content=self._config["fit_to_content"],
            flags=self._config["label_flags"],
        )

        # --------------------------------------------------------- #
        self.labelList = LabelListWidget()
        self.lastOpenDir = None
        # --------------------------------------------------------- #

        # --------------------------------------------------------- #
        self.flag_widget = QtWidgets.QListWidget()
        if config["flags"]:
            self.loadFlags({k: False for k in config["flags"]})
        self.flag_widget.itemChanged.connect(self.setDirty)
        # --------------------------------------------------------- #

        # --------------------------------------------------------- #
        self.labelList.itemSelectionChanged.connect(self.labelSelectionChanged)
        self.labelList.itemDoubleClicked.connect(self.editLabel)
        self.labelList.itemChanged.connect(self.labelItemChanged)
        self.labelList.itemDropped.connect(self.labelOrderChanged)
        # --------------------------------------------------------- #

        # --------------------------------------------------------- #
        self.uniqLabelList = UniqueLabelQListWidget()
        self.uniqLabelList.setToolTip(
            self.tr(
                "Select label to start annotating for it. "
                "Press 'Esc' to deselect."
            )
        )
        if self._config["labels"]:
            for label in self._config["labels"]:
                item = self.uniqLabelList.createItemFromLabel(label)
                self.uniqLabelList.addItem(item)
                rgb = self._get_rgb_by_label(label)
                self.uniqLabelList.setItemLabel(item, label, rgb)
        # --------------------------------------------------------- #

        # --------------------------------------------------------- #
        self.fileSearch = QtWidgets.QLineEdit()
        # self.fileSearch = QtWidgets.QHBoxLayout()
        self.fileSearch.setPlaceholderText(self.tr("搜索文件"))
        self.fileSearch.textChanged.connect(self.fileSearchChanged)
        self.fileListWidget = QtWidgets.QListWidget()
        self.fileListWidget.itemSelectionChanged.connect(
            self.fileSelectionChanged
        )
        fileListLayout = QtWidgets.QVBoxLayout()
        fileListLayout.setContentsMargins(0, 0, 0, 0)
        fileListLayout.setSpacing(0)
        fileListLayout.addWidget(self.fileSearch)
        fileListLayout.addWidget(self.fileListWidget)

        fileListWidget = QtWidgets.QWidget()
        fileListWidget.setLayout(fileListLayout)
        # --------------------------------------------------------- #

        # ---------------------添加----------------------------------------------------------------------#
        self.serverfileSearch = QtWidgets.QLineEdit()
        self.serverfileSearch.setPlaceholderText(self.tr("搜索文件"))
        self.serverfileSearch.textChanged.connect(self.fileSearchChanged)
        self.serverFileListWidget = QtWidgets.QListWidget()
        self.serverFileListWidget.itemSelectionChanged.connect(
            self.serverFileSelectionChanged
        )
        serverFileListLayout = QtWidgets.QVBoxLayout()
        serverFileListLayout.setContentsMargins(0, 0, 0, 0)
        serverFileListLayout.setSpacing(0)
        serverFileListLayout.addWidget(self.serverfileSearch)
        serverFileListLayout.addWidget(self.serverFileListWidget)
        serverFileListWidget = QtWidgets.QWidget()
        serverFileListWidget.setLayout(serverFileListLayout)

        self.currentImageORBList = ImageList()
        self.imagesList = ImageList()

        self.tqdmDialog = TqdmDialog()
        self.thread_tqdm = TestThread()
        # ---------------------end---------------------------------------------------------------------#

        self.zoomWidget = ZoomWidget()
        self.setAcceptDrops(True)

        self.canvas = self.labelList.canvas = Canvas(
            epsilon=self._config["epsilon"],
            double_click=self._config["canvas"]["double_click"],
            num_backups=self._config["canvas"]["num_backups"],
        )
        self.canvas.zoomRequest.connect(self.zoomRequest)

        scrollArea = QtWidgets.QScrollArea()
        scrollArea.setWidget(self.canvas)
        scrollArea.setWidgetResizable(True)
        self.scrollBars = {
            Qt.Vertical: scrollArea.verticalScrollBar(),
            Qt.Horizontal: scrollArea.horizontalScrollBar(),
        }
        self.canvas.scrollRequest.connect(self.scrollRequest)

        self.canvas.newShape.connect(self.newShape)
        self.canvas.shapeMoved.connect(self.setDirty)
        self.canvas.selectionChanged.connect(self.shapeSelectionChanged)
        self.canvas.drawingPolygon.connect(self.toggleDrawingSensitive)

        centralWidget = CentralWidget(self, scrollArea)
        centralWidget.addLeftLayout(fileListLayout, '本地文件列表')
        centralWidget.addLeftLayout(serverFileListLayout, '服务器文件列表')
        centralWidget.addLeftWidget(self.imagesList, '偏光图像展示')

        centralWidget.addRightWidget(self.feature, '特征填写')
        centralWidget.addRightWidget(self.labelList, '多边形标签')
        centralWidget.addRightWidget(self.uniqLabelList, '标签总统计')
        # centralWidget.addRightWidget(self.imagesList, '偏光图像展示')

        # centralWidget.addBottomWidget(self.tqdmDialog, '控制台')
        # centralWidget.addBottomWidget(self.imagesList, '偏光图像展示')
        # centralWidget.addBottomWidget(self.flag_widget, '标记')
        # centralWidget.addBottomWidget(self.feature, '特征填写')
        # centralWidget.addBottomWidget(self.currentImageORBList, '对齐图像展示')

        self.setCentralWidget(centralWidget)

        # Actions
        action = functools.partial(utils.newAction, self)
        shortcuts = self._config["shortcuts"]
        quit = action(
            self.tr("&Quit"),
            self.close,
            shortcuts["quit"],
            "quit",
            self.tr("Quit application"),
        )
        open_ = action(
            self.tr("&Open"),
            self.openFile,
            shortcuts["open"],
            "open",
            self.tr("Open image or label file"),
        )
        opendir = action(
            self.tr("&打开目录"),
            self.openDirDialog,
            shortcuts["open_dir"],
            "open",
            self.tr(u"Open Dir"),
        )
        openNextImg = action(
            self.tr("&下一张"),
            self.openNextImg,
            shortcuts["open_next"],
            "next",
            self.tr(u"Open next (hold Ctl+Shift to copy labels)"),
            enabled=False,
        )
        openPrevImg = action(
            self.tr("&上一张"),
            self.openPrevImg,
            shortcuts["open_prev"],
            "prev",
            self.tr(u"Open prev (hold Ctl+Shift to copy labels)"),
            enabled=False,
        )
        save = action(
            self.tr("&保存"),
            self.saveFile,
            shortcuts["save"],
            "save",
            self.tr("Save labels to file"),
            enabled=False,
        )
        saveAs = action(
            self.tr("&另存为"),
            self.saveFileAs,
            shortcuts["save_as"],
            "save-as",
            self.tr("Save labels to a different file"),
            enabled=False,
        )

        deleteFile = action(
            self.tr("&删除文件"),
            self.deleteFile,
            shortcuts["delete_file"],
            "delete",
            self.tr("Delete current label file"),
            enabled=False,
        )

        changeOutputDir = action(
            self.tr("&Change Output Dir"),
            slot=self.changeOutputDirDialog,
            shortcut=shortcuts["save_to"],
            icon="open",
            tip=self.tr(u"Change where annotations are loaded/saved"),
        )
        saveWithImageData = action(
            text="Save With Image Data",
            slot=self.enableSaveImageWithData,
            tip="Save image data in label file",
            checkable=False,
            checked=self._config["store_data"],
        )
        saveAuto = action(
            text=self.tr("Save &Automatically"),
            slot=lambda x: self.actions.saveAuto.setChecked(x),
            icon="save",
            tip=self.tr("Save automatically"),
            checkable=True,
            enabled=True,
        )
        saveAuto.setChecked(self._config["auto_save"])

        close = action(
            "&Close",
            self.closeFile,
            shortcuts["close"],
            "close",
            "Close current file",
        )

        toggle_keep_prev_mode = action(
            self.tr("Keep Previous Annotation"),
            self.toggleKeepPrevMode,
            shortcuts["toggle_keep_prev_mode"],
            None,
            self.tr('Toggle "keep pevious annotation" mode'),
            checkable=True,
        )
        toggle_keep_prev_mode.setChecked(self._config["keep_prev"])

        createMode = action(
            self.tr("创建多边形"),
            lambda: self.toggleDrawMode(False, createMode="polygon"),
            shortcuts["create_polygon"],
            "objects",
            self.tr("Start drawing polygons"),
            enabled=False,
        )
        createPointMode = action(
            self.tr("创建控制点"),
            lambda: self.toggleDrawMode(False, createMode="point"),
            shortcuts["create_point"],
            "objects",
            self.tr("Start drawing points"),
            enabled=False,
        )
        editMode = action(
            self.tr("编辑多边形"),
            self.setEditMode,
            shortcuts["edit_polygon"],
            "edit",
            self.tr("Move and edit the selected polygons"),
            enabled=False,
        )

        delete = action(
            self.tr("删除多边形"),
            self.deleteSelectedShape,
            shortcuts["delete_polygon"],
            "cancel",
            self.tr("Delete the selected polygons"),
            enabled=False,
        )
        copy = action(
            self.tr("复制多边形"),
            self.copySelectedShape,
            shortcuts["duplicate_polygon"],
            "copy",
            self.tr("Create a duplicate of the selected polygons"),
            enabled=False,
        )
        undoLastPoint = action(
            self.tr("撤销上一个画的点"),
            self.canvas.undoLastPoint,
            shortcuts["undo_last_point"],
            "undo",
            self.tr("Undo last drawn point"),
            enabled=False,
        )
        addPointToEdge = action(
            text=self.tr("在边上加一个控制点"),
            slot=self.canvas.addPointToEdge,
            shortcut=shortcuts["add_point_to_edge"],
            icon="edit",
            tip=self.tr("Add point to the nearest edge"),
            enabled=False,
        )
        removePoint = action(
            text="移除选中的点",
            slot=self.removeSelectedPoint,
            icon="edit",
            tip="Remove selected point from polygon",
            enabled=False,
        )

        undo = action(
            self.tr("撤销"),
            self.undoShapeEdit,
            shortcuts["undo"],
            "undo",
            self.tr("Undo last add and edit of shape"),
            enabled=False,
        )

        hideAll = action(
            self.tr("&Hide\nPolygons"),
            functools.partial(self.togglePolygons, False),
            icon="eye",
            tip=self.tr("Hide all polygons"),
            enabled=False,
        )
        showAll = action(
            self.tr("&Show\nPolygons"),
            functools.partial(self.togglePolygons, True),
            icon="eye",
            tip=self.tr("Show all polygons"),
            enabled=False,
        )

        help = action(
            self.tr("&Tutorial"),
            self.tutorial,
            icon="help",
            tip=self.tr("Show tutorial page"),
        )

        # --------------------------------------------------------------------------- #
        # 修改
        openOnline = action(
            self.tr("连接服务器"),
            self.open_online,
            icon="connect",
            tip=self.tr("Open image or label file"),
        )

        feature_ = action(
            self.tr("特征描述"),
            self.feature_w,
            icon="color-line",
            tip=self.tr("给该颗粒进行详细的特征描述"),
            enabled=False,
        )

        pimageContrast = action(
            self.tr("偏光图像对比"),
            self.Pimage_contrast,
            shortcuts["Pimage_contrast"],
            "layer-contrast",
            self.tr("展示该薄片样本的单偏光和正价偏光图像"),
            enabled=False,
        )
        # 预处理
        lightBalance = action(
            self.tr("光照均衡化"),
            self.Light_balance,
            shortcuts["Light_balance"],
            "light-balance",
            self.tr("光照均衡化"),
            enabled=False,
        )
        smoothDenoising = action(
            self.tr("平滑去噪"),
            self.Smooth_denoising,
            shortcuts["Smooth_denoising"],
            "smooth-denoising",
            self.tr("平滑去噪"),
            enabled=False,
        )
        blur = action(
            self.tr("图像模糊"),
            self.blur,
            None,
            "smooth-denoising",
            self.tr("图像模糊"),
            enabled=False,
        )
        detail = action(
            self.tr("细节增强"),
            self.detail,
            None,
            "smooth-denoising",
            self.tr("细节增强滤波"),
            enabled=False,
        )
        edge_enhance = action(
            self.tr("边缘增强"),
            self.edge_enhance,
            None,
            "smooth-denoising",
            self.tr("边缘增强滤波，突出、加强和改善图像中不同灰度区域之间的边界和轮廓的图像增强方法。"),
            enabled=False,
        )
        sharpen = action(
            self.tr("图像锐化"),
            self.sharpen,
            None,
            "smooth-denoising",
            self.tr("图像锐化"),
            enabled=False,
        )
        gassBlur = action(
            self.tr("高斯滤波"),
            self.gassBlur,
            None,
            "smooth-denoising",
            self.tr("高斯滤波"),
            enabled=False,
        )
        tranditionalSegementation = action(
            self.tr("智能分割1"),
            self.auto_segmentation,
            shortcuts["auto_segmentation"],
            "auto",
            self.tr("针对面孔率高的图片，进行智能分割"),
            enabled=False,
        )
        semanticSegmentation = action(
            self.tr("智能分割2"),
            self.auto_segmentation1,
            shortcuts["layer_contrast"],
            "auto",
            self.tr("针对常规面孔率低的图片，进行智能分割"),
            enabled=False,
        )
        autoRecognition = action(
            self.tr("智能识别"),
            self.auto_recognition,
            shortcuts["auto_recognition"],
            "recognition",
            self.tr("在智能分割的基础上，进行智能识别颗粒矿物"),
            enabled=False,
        )

        merge = action(
            self.tr('融合多边形'),
            self.merge,
            None,
            'merge',
            # self.tr('融合多边形'),
            enabled=False,
        )
        singleMerge = action(
            self.tr('单个多边形内部融合'),
            self.canvas.mergeShapeInner,
            None,
            'merge',
            # self.tr('单个多边形内部融合'),
            enabled=False,
        )
        split = action(
            self.tr('分割多边形'),
            self.canvas.splitShape,
            None,
            'split',
            # self.tr('分割多边形'),
            enabled=False,
        )
        excelPore = action(
            self.tr('生成孔隙报告'),
            self.makeExecelPore,
            None,
            'save',
            enabled=False,
        )
        excelParticl = action(
            self.tr('生成颗粒粒径报告'),
            self.makeExecelParticl,
            None,
            'save',
            enabled=False,
        )
        orbWidget = action(
            self.tr('对齐工具'),
            # functools.partial(self.orb.show, self),
            self.orb.show,
            None,
            'save',
            enabled=True,
        )
        autoORB = action(
            self.tr('自动对齐'),
            self.autoORB,
            None,
            'merge',
            enabled=False,
        )
        # --------------------------------------------------------------------------- #
        # --------------------------------------------------------------------------- #
        # --------------------------------------------------------------------------- #
        # ----------------------------------------------------------------------------------- #

        zoom = QtWidgets.QWidgetAction(self)
        zoom.setDefaultWidget(self.zoomWidget)
        self.zoomWidget.setWhatsThis(
            self.tr(
                "Zoom in or out of the image. Also accessible with "
                "{} and {} from the canvas."
            ).format(
                utils.fmtShortcut(
                    "{},{}".format(shortcuts["zoom_in"], shortcuts["zoom_out"])
                ),
                utils.fmtShortcut(self.tr("Ctrl+Wheel")),
            )
        )
        self.zoomWidget.setEnabled(False)

        zoomIn = action(
            self.tr("Zoom &In"),
            functools.partial(self.addZoom, 1.1),
            shortcuts["zoom_in"],
            "zoom-in",
            self.tr("Increase zoom level"),
            enabled=False,
        )
        zoomOut = action(
            self.tr("&Zoom Out"),
            functools.partial(self.addZoom, 0.9),
            shortcuts["zoom_out"],
            "zoom-out",
            self.tr("Decrease zoom level"),
            enabled=False,
        )
        zoomOrg = action(
            self.tr("&Original size"),
            functools.partial(self.setZoom, 100),
            shortcuts["zoom_to_original"],
            "zoom",
            self.tr("Zoom to original size"),
            enabled=False,
        )
        fitWindow = action(
            self.tr("&Fit Window"),
            self.setFitWindow,
            shortcuts["fit_window"],
            "fit-window",
            self.tr("Zoom follows window size"),
            checkable=True,
            enabled=False,
        )
        fitWidth = action(
            self.tr("Fit &Width"),
            self.setFitWidth,
            shortcuts["fit_width"],
            "fit-width",
            self.tr("Zoom follows window width"),
            checkable=True,
            enabled=False,
        )
        brightnessContrast = action(
            "亮度和对比度",
            self.brightnessContrast,
            None,
            "color",
            "调节亮度和对比度",
            enabled=False,
        )
        # Group zoom controls into a list for easier toggling.
        zoomActions = (
            self.zoomWidget,
            zoomIn,
            zoomOut,
            zoomOrg,
            fitWindow,
            fitWidth,
        )
        self.zoomMode = self.FIT_WINDOW
        fitWindow.setChecked(Qt.Checked)
        self.scalers = {
            self.FIT_WINDOW: self.scaleFitWindow,
            self.FIT_WIDTH: self.scaleFitWidth,
            # Set to one to scale to 100% when loading files.
            self.MANUAL_ZOOM: lambda: 1,
        }

        edit = action(
            self.tr("编辑标签"),
            self.editLabel,
            shortcuts["edit_label"],
            "edit",
            self.tr("Modify the label of the selected polygon"),
            enabled=False,
        )

        fill_drawing = action(
            self.tr("Fill Drawing Polygon"),
            self.canvas.setFillDrawing,
            None,
            "color",
            self.tr("Fill polygon while drawing"),
            checkable=True,
            enabled=True,
        )
        fill_drawing.trigger()

        # Lavel list context menu.
        labelMenu = QtWidgets.QMenu()
        utils.addActions(labelMenu, (edit, delete))
        self.labelList.setContextMenuPolicy(Qt.CustomContextMenu)
        self.labelList.customContextMenuRequested.connect(
            self.popLabelListMenu
        )

        # Store actions for further handling.
        self.actions = utils.struct(
            saveAuto=saveAuto,
            saveWithImageData=saveWithImageData,
            changeOutputDir=changeOutputDir,
            save=save,
            saveAs=saveAs,
            open=open_,
            close=close,
            deleteFile=deleteFile,
            toggleKeepPrevMode=toggle_keep_prev_mode,
            delete=delete,
            edit=edit,
            copy=copy,
            undoLastPoint=undoLastPoint,
            undo=undo,
            addPointToEdge=addPointToEdge,
            removePoint=removePoint,
            createMode=createMode,
            editMode=editMode,
            # createRectangleMode=createRectangleMode,
            # createCircleMode=createCircleMode,
            # createLineMode=createLineMode,
            createPointMode=createPointMode,
            # createLineStripMode=createLineStripMode,
            zoom=zoom,
            zoomIn=zoomIn,
            zoomOut=zoomOut,
            zoomOrg=zoomOrg,
            fitWindow=fitWindow,
            fitWidth=fitWidth,
            brightnessContrast=brightnessContrast,
            zoomActions=zoomActions,
            openNextImg=openNextImg,
            openPrevImg=openPrevImg,

            # --------------------------------------------------------------------------- #
            # 修改
            openOnline=openOnline,
            feature_=feature_,
            pimageContrast=pimageContrast,
            lightBalance=lightBalance,
            smoothDenoising=smoothDenoising,
            gassBlur=gassBlur,
            sharpen=sharpen,
            edge_enhance=edge_enhance,
            detail=detail,
            blur=blur,
            tranditionalSegementation=tranditionalSegementation,
            autoRecognition=autoRecognition,
            semanticSegmentation=semanticSegmentation,
            generateReport1=generateReport1,
            generateReport2=generateReport2,
            merge=merge,
            singleMerge=singleMerge,
            split=split,
            excelPore=excelPore,
            excelParticl=excelParticl,
            orbWidget=orbWidget,
            autoORB=autoORB,
            # --------------------------------------------------------------------------- #
            # --------------------------------------------------------------------------- #
            # --------------------------------------------------------------------------- #
            fileMenuActions=(open_, opendir, save, saveAs, close, quit),
            tool=(),
            # XXX: need to add some actions here to activate the shortcut
            editMenu=(
                edit,
                copy,
                delete,
                None,
                undo,
                undoLastPoint,
                None,
                addPointToEdge,
                None,
                toggle_keep_prev_mode,
            ),
            # menu shown at right click
            menu=(
                createMode,
                editMode,
                edit,
                copy,
                delete,
                undo,
                undoLastPoint,
                addPointToEdge,
                removePoint,
                singleMerge,
                split,
                merge,
            ),
            onLoadActive=(
                close,
                createMode,
                createPointMode,
                editMode,
                brightnessContrast,
                lightBalance,
                smoothDenoising,
                pimageContrast,
                autoRecognition,
                semanticSegmentation,
                tranditionalSegementation,
            ),
            onShapesPresent=(saveAs, hideAll, showAll),
        )

        self.canvas.edgeSelected.connect(self.canvasShapeEdgeSelected)
        self.canvas.vertexSelected.connect(self.actions.removePoint.setEnabled)

        self.menus = utils.struct(
            file=self.menu(self.tr("&文件")),
            edit=self.menu(self.tr("&编辑")),
            view=self.menu(self.tr("&视图")),
            preProcess=self.menu(self.tr("&图片预处理")),
            excel=self.menu(self.tr("&导出报告")),
            tools=self.menu(self.tr("&小工具")),
            help=self.menu(self.tr("&帮助")),


            recentFiles=QtWidgets.QMenu(self.tr("Open &Recent")),
            labelList=labelMenu,
        )
        utils.addActions(
            self.menus.tools,
            (
                orbWidget,
            ),
        )

        utils.addActions(
            self.menus.excel,
            (
                excelPore,
                excelParticl,
                generateReport1,
                generateReport2,
            ),
        )
        utils.addActions(
            self.menus.preProcess,
            (
                lightBalance,
                brightnessContrast,
                smoothDenoising,
                edge_enhance,
                detail,
                gassBlur,
                sharpen,
                blur

            ),
        )

        utils.addActions(
            self.menus.file,
            (
                open_,
                openNextImg,
                openPrevImg,
                opendir,
                self.menus.recentFiles,
                save,
                saveAs,
                saveAuto,
                changeOutputDir,
                saveWithImageData,
                close,
                deleteFile,
                None,
                quit,
            ),
        )
        utils.addActions(self.menus.help, (help,))
        utils.addActions(
            self.menus.view,
            (
                None,
                fill_drawing,
                None,
                hideAll,
                showAll,
                None,
                zoomIn,
                zoomOut,
                zoomOrg,
                None,
                fitWindow,
                fitWidth,
                None,
                brightnessContrast,
            ),
        )

        self.menus.file.aboutToShow.connect(self.updateFileMenu)

        # Custom context menu for the canvas widget:
        utils.addActions(self.canvas.menus[0], self.actions.menu)
        utils.addActions(
            self.canvas.menus[1],
            (
                action("&Copy here", self.copyShape),
                action("&Move here", self.moveShape),
            ),
        )

        self.tools = self.toolbar("Tools")
        # Menu buttons on Left
        self.actions.tool = (
            # open_,
            opendir,
            openOnline,
            openPrevImg,
            openNextImg,
            save,
            deleteFile,
            None,
            # feature_,
            autoORB,
            pimageContrast,
            None,
            createMode,
            editMode,
            delete,
            None,
            tranditionalSegementation,
            semanticSegmentation,
            None,
            autoRecognition,
            None,
            generateReport1,
            generateReport2,
            # fitWidth,
            # fitWindow
        )

        self.statusBar().showMessage(self.tr("%s started.") % __appname__)
        self.statusBar().show()

        if output_file is not None and self._config["auto_save"]:
            logger.warn(
                "If `auto_save` argument is True, `output_file` argument "
                "is ignored and output filename is automatically "
                "set as IMAGE_BASENAME.json."
            )
        self.output_file = output_file
        self.output_dir = output_dir

        # Application state.
        # ---------------------------------------------------------------

        self.Image_type = 0
        self.IsServer = False
        # ---------------------------------------------------------------

        self.image = QtGui.QImage()
        self.imagePath = None
        self.recentFiles = []
        self.maxRecent = 7
        self.otherData = None
        self.zoom_level = 100
        self.fit_window = False
        self.zoom_values = {}  # key=filename, value=(zoom_mode, zoom_value)
        self.brightnessContrast_values = {}
        self.scroll_values = {
            Qt.Horizontal: {},
            Qt.Vertical: {},
        }  # key=filename, value=scroll_value

        if filename is not None and osp.isdir(filename):
            self.importDirImages(filename, load=False)
        else:
            self.filename = filename

        if config["file_search"]:
            self.fileSearch.setText(config["file_search"])
            self.fileSearchChanged()

        # XXX: Could be completely declarative.
        # Restore application settings.
        self.settings = QtCore.QSettings("labelme", "labelme")
        # FIXME: QSettings.value can return None on PyQt4
        self.recentFiles = self.settings.value("recentFiles", []) or []
        size = self.settings.value("window/size", QtCore.QSize(600, 500))
        position = self.settings.value("window/position", QtCore.QPoint(0, 0))
        self.resize(size)
        self.move(position)
        # or simply:
        # self.restoreGeometry(settings['window/geometry']
        self.restoreState(
            self.settings.value("window/state", QtCore.QByteArray())
        )

        # Populate the File menu dynamically.
        self.updateFileMenu()
        # Since loading the file may take some time,
        # make sure it runs in the background.
        if self.filename is not None:
            self.queueEvent(functools.partial(self.loadFile, self.filename))

        # Callbacks:
        self.zoomWidget.valueChanged.connect(self.paintCanvas)

        self.populateModeActions()

        # self.firstStart = True
        # if self.firstStart:
        #    QWhatsThis.enterWhatsThisMode()

        # ----------------------登录---------------------- #
        self.loginDialog = LoginDialog()
        self.loginDialog.loginBtn.clicked.connect(self.login)
        if not self.isLogin:
            self.loginDialog.exec_()
        # ----------------------------------------------- #

    def login(self):
        username = self.loginDialog.username.text()
        password = self.loginDialog.password.text()
        # 发送请求
        # 如果登录成功，将 self.isLogin 置为True
        # 否则为False
        url = '/user/login'
        data = {
            'userName': username,
            'password': password
        }
        res = requests.post(self.server_path + url, json=data)
        if res.status_code == 200 and res.json()['code'] == 0:
            self.isLogin = True
            self.loginDialog.close()
        else:
            self.isLogin = False
            QMessageBox.information(self, '登录错误', '用户名或密码错误，请重新输入')

    def menu(self, title, actions=None):
        menu = self.menuBar().addMenu(title)
        if actions:
            utils.addActions(menu, actions)
        return menu

    def toolbar(self, title, actions=None):
        toolbar = ToolBar(title)
        toolbar.setObjectName("%sToolBar" % title)
        # toolbar.setOrientation(Qt.Vertical)
        toolbar.setToolButtonStyle(Qt.ToolButtonTextUnderIcon)
        if actions:
            utils.addActions(toolbar, actions)
        self.addToolBar(Qt.TopToolBarArea, toolbar)
        return toolbar

    # Support Functions

    def noShapes(self):
        return not len(self.labelList)

    def populateModeActions(self):
        tool, menu = self.actions.tool, self.actions.menu
        self.tools.clear()
        utils.addActions(self.tools, tool)
        self.canvas.menus[0].clear()
        utils.addActions(self.canvas.menus[0], menu)
        self.menus.edit.clear()
        actions = (
            self.actions.createMode,
            self.actions.editMode,
        )
        utils.addActions(self.menus.edit, actions + self.actions.editMenu)

    def setDirty(self):
        # Even if we autosave the file, we keep the ability to undo
        self.actions.undo.setEnabled(self.canvas.isShapeRestorable)

        if self._config["auto_save"] or self.actions.saveAuto.isChecked():
            label_file = osp.splitext(self.imagePath)[0] + ".json"
            if self.output_dir:
                label_file_without_path = osp.basename(label_file)
                label_file = osp.join(self.output_dir, label_file_without_path)
            self.saveLabels(label_file)
            return
        self.dirty = True
        self.actions.save.setEnabled(True)
        title = __appname__
        if self.filename is not None:
            title = "{} - {}*".format(title, self.filename)
        self.setWindowTitle(title)

    def setClean(self):
        self.dirty = False
        self.actions.save.setEnabled(False)
        self.actions.createMode.setEnabled(True)
        self.actions.createPointMode.setEnabled(True)
        # 修改
        self.actions.split.setEnabled(True)
        self.actions.singleMerge.setEnabled(True)
        self.actions.merge.setEnabled(True)
        self.actions.feature_.setEnabled(True)
        self.actions.excelParticl.setEnabled(True)
        self.actions.excelPore.setEnabled(True)
        self.actions.generateReport1.setEnabled(True)
        self.actions.generateReport2.setEnabled(True)
        self.actions.pimageContrast.setEnabled(True)
        self.actions.tranditionalSegementation.setEnabled(True)
        self.actions.semanticSegmentation.setEnabled(True)
        self.actions.autoRecognition.setEnabled(True)
        self.actions.fitWidth.setEnabled(True)
        self.actions.blur.setEnabled(True)
        self.actions.detail.setEnabled(True)
        self.actions.edge_enhance.setEnabled(True)
        self.actions.sharpen.setEnabled(True)
        self.actions.gassBlur.setEnabled(True)
        self.actions.autoORB.setEnabled(True)
        # ---------------------------------------------------------- #
        title = __appname__
        if self.filename is not None:
            title = "{} - {}".format(title, self.filename)
        self.setWindowTitle(title)

        if self.hasLabelFile():
            self.actions.deleteFile.setEnabled(True)
        else:
            self.actions.deleteFile.setEnabled(False)

    def toggleActions(self, value=True):
        """Enable/Disable widgets which depend on an opened image."""
        for z in self.actions.zoomActions:
            z.setEnabled(value)
        for action in self.actions.onLoadActive:
            action.setEnabled(value)

    def canvasShapeEdgeSelected(self, selected, shape):
        self.actions.addPointToEdge.setEnabled(
            selected and shape and shape.canAddPoint()
        )

    def queueEvent(self, function):
        QtCore.QTimer.singleShot(0, function)

    def status(self, message, delay=5000):
        self.statusBar().showMessage(message, delay)

    def resetState(self):
        self.labelList.clear()
        self.filename = None
        self.imagePath = None
        self.imageData = None
        self.labelFile = None
        self.otherData = None
        self.canvas.resetState()

    def currentItem(self):
        items = self.labelList.selectedItems()
        if items:
            return items[0]
        return None

    def addRecentFile(self, filename):
        if filename in self.recentFiles:
            self.recentFiles.remove(filename)
        elif len(self.recentFiles) >= self.maxRecent:
            self.recentFiles.pop()
        self.recentFiles.insert(0, filename)

    # Callbacks

    def undoShapeEdit(self):
        self.canvas.restoreShape()
        self.labelList.clear()
        self.loadShapes(self.canvas.shapes)
        self.actions.undo.setEnabled(self.canvas.isShapeRestorable)

    def tutorial(self):
        url = "https://www.baidu.com"
        webbrowser.open(url)

    def toggleDrawingSensitive(self, drawing=True):
        """Toggle drawing sensitive.

        In the middle of drawing, toggling between modes should be disabled.
        """
        self.actions.editMode.setEnabled(not drawing)
        self.actions.undoLastPoint.setEnabled(drawing)
        self.actions.undo.setEnabled(not drawing)
        self.actions.delete.setEnabled(not drawing)

    def toggleDrawMode(self, edit=True, createMode="polygon"):
        self.canvas.setEditing(edit)
        self.canvas.createMode = createMode
        if edit:
            self.actions.createMode.setEnabled(True)
        else:
            if createMode == "polygon":
                self.actions.createMode.setEnabled(False)
            else:
                raise ValueError("Unsupported createMode: %s" % createMode)
        self.actions.editMode.setEnabled(not edit)

    def setEditMode(self):
        self.toggleDrawMode(True)

    def updateFileMenu(self):
        current = self.filename

        def exists(filename):
            return osp.exists(str(filename))

        menu = self.menus.recentFiles
        menu.clear()
        files = [f for f in self.recentFiles if f != current and exists(f)]
        for i, f in enumerate(files):
            icon = utils.newIcon("labels")
            action = QtWidgets.QAction(
                icon, "&%d %s" % (i + 1, QtCore.QFileInfo(f).fileName()), self
            )
            action.triggered.connect(functools.partial(self.loadRecent, f))
            menu.addAction(action)

    def popLabelListMenu(self, point):
        self.menus.labelList.exec_(self.labelList.mapToGlobal(point))

    def validateLabel(self, label):
        # no validation
        if self._config["validate_label"] is None:
            return True

        for i in range(self.uniqLabelList.count()):
            label_i = self.uniqLabelList.item(i).data(Qt.UserRole)
            if self._config["validate_label"] in ["exact"]:
                if label_i == label:
                    return True
        return False

    def editLabel(self, item=None):
        if item and not isinstance(item, LabelListWidgetItem):
            raise TypeError("item must be LabelListWidgetItem type")

        if not self.canvas.editing():
            return
        if not item:
            item = self.currentItem()
        if item is None:
            return
        shape = item.shape()
        if shape is None:
            return
        text, flags, group_id = self.labelDialog.popUp(
            text=shape.label,
            flags=shape.flags,
            group_id=shape.group_id,
        )
        if text is None:
            return
        if not self.validateLabel(text):
            self.errorMessage(
                self.tr("Invalid label"),
                self.tr("Invalid label '{}' with validation type '{}'").format(
                    text, self._config["validate_label"]
                ),
            )
            return
        shape.label = text
        shape.flags = flags
        shape.group_id = group_id
        if shape.group_id is None:
            rgb = self._get_rgb_by_label(shape.label)
            r, g, b = rgb
            item.setText(
                '{} <font color="#{:02x}{:02x}{:02x}">●</font>'.format(
                    text, r, g, b
                )
            )
            shape.line_color = QtGui.QColor(r, g, b)
            shape.vertex_fill_color = QtGui.QColor(r, g, b)
            shape.fill_color = QtGui.QColor(r, g, b, 128)
            shape.select_line_color = QtGui.QColor(255, 255, 255)
            shape.select_fill_color = QtGui.QColor(r, g, b, 155)
            shape.hvertex_fill_color = QtGui.QColor(255, 255, 255)
            # item.setText(shape.label)
        else:
            item.setText("{} ({})".format(shape.label, shape.group_id))
        self.setDirty()
        if not self.uniqLabelList.findItemsByLabel(shape.label):
            item = QtWidgets.QListWidgetItem()
            item.setData(Qt.UserRole, shape.label)
            self.uniqLabelList.addItem(item)

    def fileSearchChanged(self):
        self.importDirImages(
            self.lastOpenDir,
            pattern=self.fileSearch.text(),
            load=False,
        )

    def fileSelectionChanged(self):
        items = self.fileListWidget.selectedItems()
        if not items:
            return
        item = items[0]
        if not self.mayContinue():
            return
        currIndex = self.imageList.index(str(item.text()))
        if currIndex < len(self.imageList):
            filename = self.imageList[currIndex]
            if filename:
                self.loadFile(filename)
        # -------------清空bottom展示的图片-------------------------------------
        # self.imagesList.clear()
        # self.currentImageORBList.clear()

    # React to canvas signals.
    def shapeSelectionChanged(self, selected_shapes):
        self._noSelectionSlot = True
        for shape in self.canvas.selectedShapes:
            shape.selected = False
        self.labelList.clearSelection()
        self.canvas.selectedShapes = selected_shapes
        for shape in self.canvas.selectedShapes:
            shape.selected = True
            item = self.labelList.findItemByShape(shape)
            self.labelList.selectItem(item)
            self.labelList.scrollToItem(item)
        self._noSelectionSlot = False
        n_selected = len(selected_shapes)
        self.actions.delete.setEnabled(n_selected)
        self.actions.copy.setEnabled(n_selected)
        self.actions.edit.setEnabled(n_selected == 1)

    def addLabel(self, shape):
        if shape.group_id is None:
            text = shape.label
        else:
            text = "{} ({})".format(shape.label, shape.group_id)
        label_list_item = LabelListWidgetItem(text, shape)
        self.labelList.addItem(label_list_item)
        if not self.uniqLabelList.findItemsByLabel(shape.label):
            item = self.uniqLabelList.createItemFromLabel(shape.label)
            self.uniqLabelList.addItem(item)
            rgb = self._get_rgb_by_label(shape.label)
            self.uniqLabelList.setItemLabel(item, shape.label, rgb)
        self.labelDialog.addLabelHistory(shape.label)
        for action in self.actions.onShapesPresent:
            action.setEnabled(True)

        rgb = self._get_rgb_by_label(shape.label)

        r, g, b = rgb
        label_list_item.setText(
            '{} <font color="#{:02x}{:02x}{:02x}">●</font>'.format(
                text, r, g, b
            )
        )
        shape.line_color = QtGui.QColor(r, g, b)
        shape.vertex_fill_color = QtGui.QColor(r, g, b)
        shape.hvertex_fill_color = QtGui.QColor(255, 255, 255)
        shape.fill_color = QtGui.QColor(r, g, b, 128)
        shape.select_line_color = QtGui.QColor(255, 255, 255)
        shape.select_fill_color = QtGui.QColor(r, g, b, 155)

    def _get_rgb_by_label(self, label):
        if self._config["shape_color"] == "auto":
            item = self.uniqLabelList.findItemsByLabel(label)[0]
            label_id = self.uniqLabelList.indexFromItem(item).row() + 1
            label_id += self._config["shift_auto_shape_color"]
            return LABEL_COLORMAP[label_id % len(LABEL_COLORMAP)]
        elif (
            self._config["shape_color"] == "manual"
            and self._config["label_colors"]
            and label in self._config["label_colors"]
        ):
            return self._config["label_colors"][label]
        elif self._config["default_shape_color"]:
            return self._config["default_shape_color"]

    def remLabels(self, shapes):
        for shape in shapes:
            item = self.labelList.findItemByShape(shape)
            self.labelList.removeItem(item)

    def loadShapes(self, shapes, replace=True):
        self._noSelectionSlot = True
        for shape in shapes:
            self.addLabel(shape)
        self.labelList.clearSelection()
        self._noSelectionSlot = False
        self.canvas.loadShapes(shapes, replace=replace)

    def loadLabels(self, shapes):
        s = []
        for shape in shapes:
            label = shape["label"]
            points = shape["points"]
            shape_type = shape["shape_type"]
            flags = shape["flags"]
            group_id = shape["group_id"]
            # other_data = shape["other_data"]
            features = shape["features"]
            # feature_id = shape["feature_id"]
            # face_porosity = shape["face_porosity"]
            # face_area = shape["face_area"]

            if not points:
                # skip point-empty shape
                continue

            shape = Shape(
                label=label,
                shape_type=shape_type,
                group_id=group_id,
            )
            for x, y in points:
                shape.addPoint(QtCore.QPointF(x, y))
            shape.close()

            default_flags = {}
            if self._config["label_flags"]:
                for pattern, keys in self._config["label_flags"].items():
                    if re.match(pattern, label):
                        for key in keys:
                            default_flags[key] = False
            shape.flags = default_flags
            shape.flags.update(flags)
            # shape.other_data = other_data
            shape.features = features
            # shape.feature_id = feature_id
            # shape.face_porosity = face_porosity
            # shape.face_area = face_area
            s.append(shape)
        self.loadShapes(s)

    def loadFlags(self, flags):
        self.flag_widget.clear()
        for key, flag in flags.items():
            item = QtWidgets.QListWidgetItem(key)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked if flag else Qt.Unchecked)
            self.flag_widget.addItem(item)

    def saveLabels(self, filename):
        self.countProps()  # 更新颗粒属性
        lf = LabelFile()

        def format_shape(s):
            data = s.other_data.copy()
            data.update(
                dict(
                    label=s.label.encode("utf-8") if PY2 else s.label,
                    points=[(p.x(), p.y()) for p in s.points],
                    group_id=s.group_id,
                    shape_type=s.shape_type,
                    flags=s.flags,
                    feature_id=s.feature_id,
                    face_porosity=s.face_porosity,
                    face_area=s.face_area,
                    features=s.features,
                )
            )
            return data

        shapes = [format_shape(item.shape()) for item in self.labelList]
        flags = {}
        for i in range(self.flag_widget.count()):
            item = self.flag_widget.item(i)
            key = item.text()
            flag = item.checkState() == Qt.Checked
            flags[key] = flag
        try:
            imagePath = osp.relpath(self.imagePath, osp.dirname(filename))
            # imageData = self.imageData if self._config["store_data"] else None
            if osp.dirname(filename) and not osp.exists(osp.dirname(filename)):
                os.makedirs(osp.dirname(filename))
            lf.save(
                filename=filename,
                shapes=shapes,
                imagePath=imagePath,
                # imageData=imageData,
                imageHeight=self.image.height(),
                imageWidth=self.image.width(),
                otherData=self.otherData,
                flags=flags,
            )
            self.labelFile = lf
            items = self.fileListWidget.findItems(
                self.imagePath, Qt.MatchExactly
            )
            if len(items) > 0:
                if len(items) != 1:
                    raise RuntimeError("There are duplicate files.")
                items[0].setCheckState(Qt.Checked)
            # disable allows next and previous image to proceed
            # self.filename = filename
            return True
        except LabelFileError as e:
            self.errorMessage(
                self.tr("Error saving label data"), self.tr("<b>%s</b>") % e
            )
            return False

    def copySelectedShape(self):
        added_shapes = self.canvas.copySelectedShapes()
        self.labelList.clearSelection()
        for shape in added_shapes:
            self.addLabel(shape)
        self.setDirty()

    def labelSelectionChanged(self):
        # --------------------------------------------------------------------------- #
        #  修改
        shapes = self.canvas.selectedShapes
        if shapes is not None and len(shapes) > 0:
            shape = shapes[0]
            self.feature.loadDataFromShape(shape)
        # --------------------------------------------------------------------------- #

        if self._noSelectionSlot:
            return
        if self.canvas.editing():
            selected_shapes = []
            for item in self.labelList.selectedItems():
                selected_shapes.append(item.shape())
            if selected_shapes:
                self.canvas.selectShapes(selected_shapes)
            else:
                self.canvas.deSelectShape()

    def labelItemChanged(self, item):
        shape = item.shape()
        self.canvas.setShapeVisible(shape, item.checkState() == Qt.Checked)

    def labelOrderChanged(self):
        self.setDirty()
        self.canvas.loadShapes([item.shape() for item in self.labelList])

    # Callback functions:
    def newShape(self):
        items = self.uniqLabelList.selectedItems()
        text = '默认名'
        if items:
            text = items[0].data(Qt.UserRole)
        flags = {}
        group_id = None
        if text:
            self.labelList.clearSelection()
            shape = self.canvas.setLastLabel(text, flags)
            shape.group_id = group_id

            shape.features = self.pelletFeatures
            shape.feature_id = len(self.canvas.shapes)
            shape.features['labelName'] = text
            self.addLabel(shape)
            # 创建新多边形时，计算其颗粒属性
            count = [[[round(point.x()), round(point.y())]] for point in shape.points]
            count = np.array(count)
            shape.face_porosity = self.lengthDiameter(count)
            shape.face_area = self.cnt_area(count)

            self.actions.editMode.setEnabled(True)
            self.actions.undoLastPoint.setEnabled(False)
            self.actions.undo.setEnabled(True)
            self.setDirty()
        else:
            self.canvas.undoLastLine()
            self.canvas.shapesBackups.pop()

    # def newShape(self):
    #     """Pop-up and give focus to the label editor.
    #
    #             position MUST be in global coordinates.
    #             """
    #     items = self.uniqLabelList.selectedItems()
    #     text = None
    #     if items:
    #         text = items[0].data(Qt.UserRole)
    #     flags = {}
    #     group_id = None
    #     if self._config["display_label_popup"] or not text:
    #         previous_text = self.labelDialog.edit.text()
    #         text, flags, group_id = self.labelDialog.popUp(text)
    #         if not text:
    #             self.labelDialog.edit.setText(previous_text)
    #
    #     if text and not self.validateLabel(text):
    #         self.errorMessage(
    #             self.tr("Invalid label"),
    #             self.tr("Invalid label '{}' with validation type '{}'").format(
    #                 text, self._config["validate_label"]
    #             ),
    #         )
    #         text = ""
    #     if text:
    #         self.labelList.clearSelection()
    #         shape = self.canvas.setLastLabel(text, flags)
    #         shape.group_id = group_id
    #         shape.features = self.pelletFeatures
    #         shape.feature_id = len(self.canvas.shapes)
    #         shape.features['labelName'] = text
    #         self.addLabel(shape)
    #
    #         # 创建新多边形时，计算其颗粒属性
    #         count = [[[round(point.x()), round(point.y())]] for point in shape.points]
    #         count = np.array(count)
    #         shape.features['facePorosity'] = self.lengthDiameter(count)
    #         shape.features['faceArea'] = self.cnt_area(count)
    #
    #         self.actions.editMode.setEnabled(True)
    #         self.actions.undoLastPoint.setEnabled(False)
    #         self.actions.undo.setEnabled(True)
    #         self.setDirty()
    #     else:
    #         self.canvas.undoLastLine()
    #         self.canvas.shapesBackups.pop()

    def scrollRequest(self, delta, orientation):
        units = -delta * 0.1  # natural scroll
        bar = self.scrollBars[orientation]
        value = bar.value() + bar.singleStep() * units
        self.setScroll(orientation, value)

    def setScroll(self, orientation, value):
        self.scrollBars[orientation].setValue(value)
        self.scroll_values[orientation][self.filename] = value

    def setZoom(self, value):
        self.actions.fitWidth.setChecked(False)
        self.actions.fitWindow.setChecked(False)
        self.zoomMode = self.MANUAL_ZOOM
        self.zoomWidget.setValue(value)
        self.zoom_values[self.filename] = (self.zoomMode, value)

    def addZoom(self, increment=1.1):
        zoom_value = self.zoomWidget.value() * increment
        if increment > 1:
            zoom_value = math.ceil(zoom_value)
        else:
            zoom_value = math.floor(zoom_value)
        self.setZoom(zoom_value)

    def zoomRequest(self, delta, pos):
        canvas_width_old = self.canvas.width()
        units = 1.1
        if delta < 0:
            units = 0.9
        self.addZoom(units)

        canvas_width_new = self.canvas.width()
        if canvas_width_old != canvas_width_new:
            canvas_scale_factor = canvas_width_new / canvas_width_old

            x_shift = round(pos.x() * canvas_scale_factor) - pos.x()
            y_shift = round(pos.y() * canvas_scale_factor) - pos.y()

            self.setScroll(
                Qt.Horizontal,
                self.scrollBars[Qt.Horizontal].value() + x_shift,
            )
            self.setScroll(
                Qt.Vertical,
                self.scrollBars[Qt.Vertical].value() + y_shift,
            )

    def setFitWindow(self, value=True):
        if value:
            self.actions.fitWidth.setChecked(False)
        self.zoomMode = self.FIT_WINDOW if value else self.MANUAL_ZOOM
        self.adjustScale()

    def setFitWidth(self, value=True):
        if value:
            self.actions.fitWindow.setChecked(False)
        self.zoomMode = self.FIT_WIDTH if value else self.MANUAL_ZOOM
        self.adjustScale()

    def onNewBrightnessContrast(self, qimage):
        self.canvas.loadPixmap(
            QtGui.QPixmap.fromImage(qimage), clear_shapes=False
        )

    def brightnessContrast(self, value):
        dialog = BrightnessContrastDialog(
            utils.img_data_to_pil(self.imageData),
            self.onNewBrightnessContrast,
            self.IsServer,
            self.filename,
            parent=self,
        )
        brightness, contrast = self.brightnessContrast_values.get(
            self.filename, (None, None)
        )
        if brightness is not None:
            dialog.slider_brightness.setValue(brightness)
        if contrast is not None:
            dialog.slider_contrast.setValue(contrast)
        dialog.exec_()

        brightness = dialog.slider_brightness.value()
        contrast = dialog.slider_contrast.value()
        self.brightnessContrast_values[self.filename] = (brightness, contrast)

    def togglePolygons(self, value):
        for item in self.labelList:
            item.setCheckState(Qt.Checked if value else Qt.Unchecked)

    def loadFile(self, filename=None):
        """Load the specified file, or the last opened file if None."""
        # changing fileListWidget loads file
        if filename in self.imageList and (
            self.fileListWidget.currentRow() != self.imageList.index(filename)
        ):
            self.fileListWidget.setCurrentRow(self.imageList.index(filename))
            self.fileListWidget.repaint()
            return

        self.resetState()
        self.canvas.setEnabled(False)
        if filename is None:
            filename = self.settings.value("filename", "")
        filename = str(filename)
        if not QtCore.QFile.exists(filename):
            self.errorMessage(
                self.tr("Error opening file"),
                self.tr("No such file: <b>%s</b>") % filename,
            )
            return False
        # assumes same name, but json extension
        self.status(self.tr("Loading %s...") % osp.basename(str(filename)))
        label_file = osp.splitext(filename)[0] + ".json"
        label_file_final = '_'.join(osp.splitext(filename)[0].split('_')[0:-3]) + ".json"
        if self.output_dir:
            label_file_without_path = osp.basename(label_file)
            label_file = osp.join(self.output_dir, label_file_without_path)
        if QtCore.QFile.exists(label_file_final) and LabelFile.is_label_file(
            label_file_final
        ):
        # if QtCore.QFile.exists(label_file) and LabelFile.is_label_file(
        #     label_file
        # ):
            try:
                self.labelFile = LabelFile(label_file)
            except LabelFileError as e:
                self.errorMessage(
                    self.tr("Error opening file"),
                    self.tr(
                        "<p><b>%s</b></p>"
                        "<p>Make sure <i>%s</i> is a valid label file."
                    )
                    % (e, label_file),
                )
                self.status(self.tr("Error reading %s") % label_file)
                return False
            self.imageData = self.labelFile.imageData
            self.imagePath = osp.join(
                osp.dirname(label_file),
                self.labelFile.imagePath,
            )
            self.otherData = self.labelFile.otherData
        else:
            self.imageData = LabelFile.load_image_file(filename)
            if self.imageData:
                self.imagePath = filename
            self.labelFile = None
        image = QtGui.QImage.fromData(self.imageData)

        if image.isNull():
            formats = [
                "*.{}".format(fmt.data().decode())
                for fmt in QtGui.QImageReader.supportedImageFormats()
            ]
            self.errorMessage(
                self.tr("Error opening file"),
                self.tr(
                    "<p>Make sure <i>{0}</i> is a valid image file.<br/>"
                    "Supported image formats: {1}</p>"
                ).format(filename, ",".join(formats)),
            )
            self.status(self.tr("Error reading %s") % filename)
            return False
        self.image = image
        self.filename = filename
        if self._config["keep_prev"]:
            prev_shapes = self.canvas.shapes
        self.canvas.loadPixmap(QtGui.QPixmap.fromImage(image))
        flags = {k: False for k in self._config["flags"] or []}
        if self.labelFile:
            self.loadLabels(self.labelFile.shapes)
            if self.labelFile.flags is not None:
                flags.update(self.labelFile.flags)
        self.loadFlags(flags)
        if self._config["keep_prev"] and self.noShapes():
            self.loadShapes(prev_shapes, replace=False)
            self.setDirty()
        else:
            self.setClean()
        self.canvas.setEnabled(True)
        # set zoom values
        is_initial_load = not self.zoom_values
        if self.filename in self.zoom_values:
            self.zoomMode = self.zoom_values[self.filename][0]
            self.setZoom(self.zoom_values[self.filename][1])
        elif is_initial_load or not self._config["keep_prev_scale"]:
            self.adjustScale(initial=True)
        # set scroll values
        for orientation in self.scroll_values:
            if self.filename in self.scroll_values[orientation]:
                self.setScroll(
                    orientation, self.scroll_values[orientation][self.filename]
                )
        # set brightness constrast values
        dialog = BrightnessContrastDialog(
            utils.img_data_to_pil(self.imageData),
            self.onNewBrightnessContrast,
            self.IsServer,
            self.filename,
            parent=self,
        )
        brightness, contrast = self.brightnessContrast_values.get(
            self.filename, (None, None)
        )
        if self._config["keep_prev_brightness"] and self.recentFiles:
            brightness, _ = self.brightnessContrast_values.get(
                self.recentFiles[0], (None, None)
            )
        if self._config["keep_prev_contrast"] and self.recentFiles:
            _, contrast = self.brightnessContrast_values.get(
                self.recentFiles[0], (None, None)
            )
        if brightness is not None:
            dialog.slider_brightness.setValue(brightness)
        if contrast is not None:
            dialog.slider_contrast.setValue(contrast)
        self.brightnessContrast_values[self.filename] = (brightness, contrast)
        if brightness is not None or contrast is not None:
            dialog.onNewValue(None)
        self.paintCanvas()
        self.addRecentFile(self.filename)
        self.toggleActions(True)
        self.canvas.setFocus()
        self.status(self.tr("Loaded %s") % osp.basename(str(filename)))
        return True

    def resizeEvent(self, event):
        if (
            self.canvas
            and not self.image.isNull()
            and self.zoomMode != self.MANUAL_ZOOM
        ):
            self.adjustScale()
        super(MainWindow, self).resizeEvent(event)

    def paintCanvas(self):
        assert not self.image.isNull(), "cannot paint null image"
        self.canvas.scale = 0.01 * self.zoomWidget.value()
        self.canvas.adjustSize()
        self.canvas.update()

    def adjustScale(self, initial=False):
        value = self.scalers[self.FIT_WINDOW if initial else self.zoomMode]()
        value = int(100 * value)
        self.zoomWidget.setValue(value)
        self.zoom_values[self.filename] = (self.zoomMode, value)

    def scaleFitWindow(self):
        """Figure out the size of the pixmap to fit the main widget."""
        e = 2.0  # So that no scrollbars are generated.
        w1 = self.centralWidget().width() - e
        h1 = self.centralWidget().height() - e
        a1 = w1 / h1
        # Calculate a new scale value based on the pixmap's aspect ratio.
        w2 = self.canvas.pixmap.width() - 0.0
        h2 = self.canvas.pixmap.height() - 0.0
        a2 = w2 / h2
        return w1 / w2 if a2 >= a1 else h1 / h2

    def scaleFitWidth(self):
        # The epsilon does not seem to work too well here.
        w = self.centralWidget().width() - 2.0
        return w / self.canvas.pixmap.width()

    def enableSaveImageWithData(self, enabled):
        self._config["store_data"] = enabled
        self.actions.saveWithImageData.setChecked(enabled)

    def closeEvent(self, event):
        if not self.mayContinue():
            event.ignore()
        self.settings.setValue(
            "filename", self.filename if self.filename else ""
        )
        self.settings.setValue("window/size", self.size())
        self.settings.setValue("window/position", self.pos())
        self.settings.setValue("window/state", self.saveState())
        self.settings.setValue("recentFiles", self.recentFiles)
        # ask the use for where to save the labels
        # self.settings.setValue('window/geometry', self.saveGeometry())

    def dragEnterEvent(self, event):
        extensions = [
            ".%s" % fmt.data().decode().lower()
            for fmt in QtGui.QImageReader.supportedImageFormats()
        ]
        if event.mimeData().hasUrls():
            items = [i.toLocalFile() for i in event.mimeData().urls()]
            if any([i.lower().endswith(tuple(extensions)) for i in items]):
                event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        if not self.mayContinue():
            event.ignore()
            return
        items = [i.toLocalFile() for i in event.mimeData().urls()]
        self.importDroppedImageFiles(items)

    # User Dialogs #

    def loadRecent(self, filename):
        if self.mayContinue():
            self.loadFile(filename)

    def openPrevImg(self, _value=False):
        self.preImage = None
        keep_prev = self._config["keep_prev"]
        if Qt.KeyboardModifiers() == (Qt.ControlModifier | Qt.ShiftModifier):
            self._config["keep_prev"] = True

        if not self.mayContinue():
            return

        if len(self.imageList) <= 0:
            return

        if self.filename is None:
            return

        currIndex = self.imageList.index(self.filename)
        if currIndex - 1 >= 0:
            filename = self.imageList[currIndex - 1]
            if filename:
                self.loadFile(filename)

        self._config["keep_prev"] = keep_prev

    def openNextImg(self, _value=False, load=True):
        self.preImage = None
        keep_prev = self._config["keep_prev"]
        if Qt.KeyboardModifiers() == (Qt.ControlModifier | Qt.ShiftModifier):
            self._config["keep_prev"] = True

        if not self.mayContinue():
            return

        if len(self.imageList) <= 0:
            return

        filename = None
        if self.filename is None:
            filename = self.imageList[0]
        else:
            currIndex = self.imageList.index(self.filename)
            if currIndex + 1 < len(self.imageList):
                filename = self.imageList[currIndex + 1]
            else:
                filename = self.imageList[-1]
        self.filename = filename

        if self.filename and load:
            self.loadFile(self.filename)

        self._config["keep_prev"] = keep_prev

    def openFile(self, _value=False):
        self.IsServer = False
        if not self.mayContinue():
            return
        path = osp.dirname(str(self.filename)) if self.filename else "."
        formats = [
            "*.{}".format(fmt.data().decode())
            for fmt in QtGui.QImageReader.supportedImageFormats()
        ]
        filters = self.tr("Image & Label files (%s)") % " ".join(
            formats + ["*%s" % LabelFile.suffix]
        )
        filename = QtWidgets.QFileDialog.getOpenFileName(
            self,
            self.tr("%s - Choose Image or Label file") % __appname__,
            path,
            filters,
        )
        if QT5:
            filename, _ = filename
        filename = str(filename)
        if filename:
            self.loadFile(filename)

    def changeOutputDirDialog(self, _value=False):
        default_output_dir = self.output_dir
        if default_output_dir is None and self.filename:
            default_output_dir = osp.dirname(self.filename)
        if default_output_dir is None:
            default_output_dir = self.currentPath()

        output_dir = QtWidgets.QFileDialog.getExistingDirectory(
            self,
            self.tr("%s - Save/Load Annotations in Directory") % __appname__,
            default_output_dir,
            QtWidgets.QFileDialog.ShowDirsOnly
            | QtWidgets.QFileDialog.DontResolveSymlinks,
        )
        output_dir = str(output_dir)

        if not output_dir:
            return

        self.output_dir = output_dir

        self.statusBar().showMessage(
            self.tr("%s . Annotations will be saved/loaded in %s")
            % ("Change Annotations Dir", self.output_dir)
        )
        self.statusBar().show()

        current_filename = self.filename
        self.importDirImages(self.lastOpenDir, load=False)

        if current_filename in self.imageList:
            # retain currently selected file
            self.fileListWidget.setCurrentRow(
                self.imageList.index(current_filename)
            )
            self.fileListWidget.repaint()

    def saveFile(self, _value=False):
        assert not self.image.isNull(), "cannot save empty image"
        if self.labelFile:
            # DL20180323 - overwrite when in directory
            self._saveFile(self.labelFile.filename)
        elif self.output_file:
            self._saveFile(self.output_file)
            self.close()
        else:
            if not self.IsServer:
                self._saveFile(self.saveFileDialog())
            else:
                self._saveFile('icons/source.json')
        if self.IsServer:
            url = '/DataManagementSystem/OriPicAdd'
            files = {
                'file': (''.join(self.filename.split('.')[:-1]) + '.json', open(self.imagePath, 'rb')),
            }
            data = {
                'typeId': 3,
            }
            requests.post(url=self.server_path + url, data=data, files=files)
            # self.Image_type = 0
            # os.remove('icons/source.json')

    def saveFileAs(self, _value=False):
        assert not self.image.isNull(), "cannot save empty image"
        self._saveFile(self.saveFileDialog())

    def saveFileDialog(self):
        caption = self.tr("%s - Choose File") % __appname__
        filters = self.tr("Label files (*%s)") % LabelFile.suffix
        if self.output_dir:
            dlg = QtWidgets.QFileDialog(
                self, caption, self.output_dir, filters
            )
        else:
            dlg = QtWidgets.QFileDialog(
                self, caption, self.currentPath(), filters
            )
        dlg.setDefaultSuffix(LabelFile.suffix[1:])
        dlg.setAcceptMode(QtWidgets.QFileDialog.AcceptSave)
        dlg.setOption(QtWidgets.QFileDialog.DontConfirmOverwrite, False)
        dlg.setOption(QtWidgets.QFileDialog.DontUseNativeDialog, False)
        basename = osp.basename(osp.splitext(self.filename)[0])
        if self.output_dir:
            default_labelfile_name = osp.join(
                self.output_dir, basename + LabelFile.suffix
            )
        else:
            default_labelfile_name = osp.join(
                self.currentPath(), basename + LabelFile.suffix
            )
        filename = dlg.getSaveFileName(
            self,
            self.tr("Choose File"),
            default_labelfile_name,
            self.tr("Label files (*%s)") % LabelFile.suffix,
        )
        if isinstance(filename, tuple):
            filename, _ = filename
        return filename

    def _saveFile(self, filename):
        self.countProps()  # 重新计算，并更新颗粒信息
        if filename and self.saveLabels(filename):
            self.addRecentFile(filename)
            self.setClean()

    def closeFile(self, _value=False):
        if not self.mayContinue():
            return
        self.resetState()
        self.setClean()
        self.toggleActions(False)
        self.canvas.setEnabled(False)
        self.actions.saveAs.setEnabled(False)

    def getLabelFile(self):
        if self.filename.lower().endswith(".json"):
            # label_file = self.filename
            # label_file = label_file
            pre = '_'.join(self.filename.split('_')[:-3])
            label_file = pre + ".json"
        else:
            pre = '_'.join(self.filename.split('_')[:-3])
            label_file = pre + ".json"
            # label_file = osp.splitext(self.filename)[0] + ".json"
        return label_file

    def deleteFile(self):
        mb = QtWidgets.QMessageBox
        msg = self.tr(
            "You are about to permanently delete this label file, "
            "proceed anyway?"
        )
        answer = mb.warning(self, self.tr("Attention"), msg, mb.Yes | mb.No)
        if answer != mb.Yes:
            return

        label_file = self.getLabelFile()
        if osp.exists(label_file):
            os.remove(label_file)
            logger.info("Label file is removed: {}".format(label_file))

            item = self.fileListWidget.currentItem()
            item.setCheckState(Qt.Unchecked)

            self.resetState()

    # Message Dialogs. #
    def hasLabels(self):
        if self.noShapes():
            self.errorMessage(
                "No objects labeled",
                "You must label at least one object to save the file.",
            )
            return False
        return True

    def hasLabelFile(self):
        if self.filename is None:
            return False

        label_file = self.getLabelFile()
        return osp.exists(label_file)

    def mayContinue(self):
        if not self.dirty:
            return True
        mb = QtWidgets.QMessageBox
        msg = self.tr('Save annotations to "{}" before closing?').format(
            self.filename
        )
        answer = mb.question(
            self,
            self.tr("Save annotations?"),
            msg,
            mb.Save | mb.Discard | mb.Cancel,
            mb.Save,
        )
        if answer == mb.Discard:
            return True
        elif answer == mb.Save:
            self.saveFile()
            return True
        else:  # answer == mb.Cancel
            return False

    def errorMessage(self, title, message):
        return QtWidgets.QMessageBox.critical(
            self, title, "<p><b>%s</b></p>%s" % (title, message)
        )

    def currentPath(self):
        return osp.dirname(str(self.filename)) if self.filename else "."

    def toggleKeepPrevMode(self):
        self._config["keep_prev"] = not self._config["keep_prev"]

    def removeSelectedPoint(self):
        self.canvas.removeSelectedPoint()
        if not self.canvas.hShape.points:
            self.canvas.deleteShape(self.canvas.hShape)
            self.remLabels([self.canvas.hShape])
            self.setDirty()
            if self.noShapes():
                for action in self.actions.onShapesPresent:
                    action.setEnabled(False)

    def deleteSelectedShape(self):
        yes, no = QtWidgets.QMessageBox.Yes, QtWidgets.QMessageBox.No
        msg = self.tr(
            "You are about to permanently delete {} polygons, "
            "proceed anyway?"
        ).format(len(self.canvas.selectedShapes))
        if yes == QtWidgets.QMessageBox.warning(
            self, self.tr("Attention"), msg, yes | no, yes
        ):
            self.remLabels(self.canvas.deleteSelected())
            self.setDirty()
            if self.noShapes():
                for action in self.actions.onShapesPresent:
                    action.setEnabled(False)

    def copyShape(self):
        self.canvas.endMove(copy=True)
        self.labelList.clearSelection()
        for shape in self.canvas.selectedShapes:
            self.addLabel(shape)
        self.setDirty()

    def moveShape(self):
        self.canvas.endMove(copy=False)
        self.setDirty()

    def openDirDialog(self, _value=False, dirpath=None):
        self.IsServer = False
        if not self.mayContinue():
            return
        defaultOpenDirPath = dirpath if dirpath else "."
        if self.lastOpenDir and osp.exists(self.lastOpenDir):
            defaultOpenDirPath = self.lastOpenDir
        else:
            defaultOpenDirPath = (osp.dirname(self.filename) if self.filename else ".")
        targetDirPath = str(
            QtWidgets.QFileDialog.getExistingDirectory(
                self,
                self.tr("%s - Open Directory") % __appname__,
                defaultOpenDirPath,
                QtWidgets.QFileDialog.ShowDirsOnly
                | QtWidgets.QFileDialog.DontResolveSymlinks,
            )
        )
        self.importDirImages(targetDirPath)

    @property
    def imageList(self):
        lst = []
        for i in range(self.fileListWidget.count()):
            item = self.fileListWidget.item(i)
            lst.append(item.text())
        return lst

    def importDroppedImageFiles(self, imageFiles):
        extensions = [
            ".%s" % fmt.data().decode().lower()
            for fmt in QtGui.QImageReader.supportedImageFormats()
        ]

        self.filename = None
        for file in imageFiles:
            if file in self.imageList or not file.lower().endswith(
                tuple(extensions)
            ):
                continue
            label_file = osp.splitext(file)[0] + ".json"
            if self.output_dir:
                label_file_without_path = osp.basename(label_file)
                label_file = osp.join(self.output_dir, label_file_without_path)
            item = QtWidgets.QListWidgetItem(file)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            if QtCore.QFile.exists(label_file) and LabelFile.is_label_file(
                label_file
            ):
                item.setCheckState(Qt.Checked)
            else:
                item.setCheckState(Qt.Unchecked)
            self.fileListWidget.addItem(item)

        if len(self.imageList) > 1:
            self.actions.openNextImg.setEnabled(True)
            self.actions.openPrevImg.setEnabled(True)

        self.openNextImg()

    def importDirImages(self, dirpath, pattern=None, load=True):
        self.actions.openNextImg.setEnabled(True)
        self.actions.openPrevImg.setEnabled(True)

        if not self.mayContinue() or not dirpath:
            return

        self.lastOpenDir = dirpath
        self.filename = None
        self.fileListWidget.clear()
        for filename in self.scanAllImages(dirpath):
            if pattern and pattern not in filename:
                continue
            label_file = osp.splitext(filename)[0] + ".json"
            if self.output_dir:
                label_file_without_path = osp.basename(label_file)
                label_file = osp.join(self.output_dir, label_file_without_path)
            label_file_final = '_'.join(osp.splitext(filename)[0].split('_')[0:-3]) + ".json"
            item = QtWidgets.QListWidgetItem(filename)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            if QtCore.QFile.exists(label_file_final) and LabelFile.is_label_file(
                label_file_final
            ):
            # if QtCore.QFile.exists(label_file) and LabelFile.is_label_file(
            #     label_file
            # ):
                item.setCheckState(Qt.Checked)
            else:
                item.setCheckState(Qt.Unchecked)
            self.fileListWidget.addItem(item)
        self.openNextImg(load=load)

    def scanAllImages(self, folderPath):
        extensions = [
            ".%s" % fmt.data().decode().lower()
            for fmt in QtGui.QImageReader.supportedImageFormats()
        ]

        images = []
        for root, dirs, files in os.walk(folderPath):
            for file in files:
                if file.lower().endswith(tuple(extensions)):
                    relativePath = osp.join(root, file)
                    images.append(relativePath)
        images.sort(key=lambda x: x.lower())

        return images

# --------------------------------------------------------------------------- #
    #  修改
    # 下载json
    def download_json(self, json_name):
        url = '/image/分割图像/'
        requests.get(self.server_path + url + json_name)

    # 文件上传
    # 1. 原图  2. 预处理图片 3. 分割图像 4. 识别结果图像 5. 样本标注图像 6. 鉴定结果
    def upload(self, img_path, type_id):
        filename = self.serverFileListWidget.selectedItems()[0].text()
        url = '/DataManagementSystem/OriPicAdd'
        files = {
            'file': (filename, open(img_path, 'rb')),
        }
        data = {
            'typeId': type_id,
        }
        res = requests.post(
            url=self.server_path + url,
            data=data,
            files=files
        )

    # 打开服务器文件
    def open_online(self):
        # if not self.isLogin:
        #     self.loginDialog.exec_()
        # else:
        #     thread = Thread(target=self.open_online_thread)
        #     thread.start()
        pass

    def open_online_thread(self):
        self.IsServer = True
        self.serverFileListWidget.clear()
        url = '/DataManagementSystem/selectByStatus'
        res = requests.post(
            url=self.server_path + url,
            data={
                'typeName': 1
            }
        )
        if res.status_code != 200:
            QMessageBox.information(self, '错误', '服务器连接失败')
            return
        for img in res.json()['data']:
            self.serverFileListWidget.addItem(img['picPath'].split('/')[-1])

    # 服务器切换图片
    def serverFileSelectionChanged(self):
        self.IsServer = True
        self.preImage = None
        if os.path.exists('icons/source.json'):
            os.remove('icons/source.json')
        # 加载原图
        filename = self.serverFileListWidget.selectedItems()[0].text()
        url = '/image/原始图像/' + filename
        res = requests.get(
            self.server_path + url,
            # params={
            #     'fileName': filename,
            #     'typeId': 1
            # }
        )
        with open('temp/contrast/' + filename, 'wb') as f:
            f.write(res.content)
        self.imagePath = os.path.abspath('temp/contrast/' + filename)
        self.loadFile(self.imagePath)
        self.filename = filename

        # 加载标签文件
        url = '/DataManagementSystem/isExit'
        filename = ''.join(filename.split('.')[:-1]) + '.json'
        res = requests.post(
            url=self.server_path+url,
            data={
                'fileName': filename,
                'typeId': 3
            }
        )
        if res.json():
            url = '/image/分割图像/'+filename
            res = requests.get(
                self.server_path + url,
            )
            self.loadLabels(res.json()["shapes"])

    # 特征填写
    def feature_w(self):
        if self.canvas.selectedShapes:
            shape = self.canvas.selectedShapes[0]
            self.feature.loadDataFromShape(shape)

    # 对齐
    def sift_kp(self, image):
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        sift = cv2.xfeatures2d_SIFT.create()
        kp, des = sift.detectAndCompute(gray, None)
        return kp, des

    def get_good_match(self, des1, des2):
        bf = cv2.BFMatcher()
        matches = bf.knnMatch(des1, des2, k=2)
        good = []
        for m, n in matches:
            if m.distance < 0.75 * n.distance:
                good.append(m)
        return good

    def siftImageAlignment(self, img1, img2):
        kp1, des1 = self.sift_kp(img1)
        kp2, des2 = self.sift_kp(img2)
        goodMatch = self.get_good_match(des1, des2)
        if len(goodMatch) > 4:
            ptsA = np.float32([kp1[m.queryIdx].pt for m in goodMatch]).reshape(-1, 1, 2)
            ptsB = np.float32([kp2[m.trainIdx].pt for m in goodMatch]).reshape(-1, 1, 2)
            H, status = cv2.findHomography(ptsA, ptsB, cv2.RANSAC, 4)
            imgOut = cv2.warpPerspective(img2, H, (img1.shape[1], img1.shape[0]),
                                         flags=cv2.INTER_LINEAR + cv2.WARP_INVERSE_MAP)
        return imgOut

    def siftImageAlignment1(self, im0, im1, im2, im3, im4, im5):
        size = (im0.shape[1], im0.shape[0])
        size1 = (2000, 2000)
        im = cv2.resize(im0, size1, interpolation=cv2.INTER_LINEAR)
        im1 = cv2.resize(im1, size1, interpolation=cv2.INTER_LINEAR)
        im2 = cv2.resize(im2, size1, interpolation=cv2.INTER_LINEAR)
        im3 = cv2.resize(im3, size1, interpolation=cv2.INTER_LINEAR)
        im4 = cv2.resize(im4, size1, interpolation=cv2.INTER_LINEAR)
        im5 = cv2.resize(im5, size1, interpolation=cv2.INTER_LINEAR)
        kp1, des1 = self.sift_kp(im)
        kp2, des2 = self.sift_kp(im1)
        goodMatch = self.get_good_match(des1, des2)
        if len(goodMatch) > 4:
            ptsA = np.float32([kp1[m.queryIdx].pt for m in goodMatch]).reshape(-1, 1, 2)
            ptsB = np.float32([kp2[m.trainIdx].pt for m in goodMatch]).reshape(-1, 1, 2)
            H, status = cv2.findHomography(ptsA, ptsB, cv2.RANSAC, 4)
            im1 = cv2.warpPerspective(im1, H, size1, flags=cv2.INTER_LINEAR + cv2.WARP_INVERSE_MAP)
            im2 = cv2.warpPerspective(im2, H, size1, flags=cv2.INTER_LINEAR + cv2.WARP_INVERSE_MAP)
            im3 = cv2.warpPerspective(im3, H, size1, flags=cv2.INTER_LINEAR + cv2.WARP_INVERSE_MAP)
            im4 = cv2.warpPerspective(im4, H, size1, flags=cv2.INTER_LINEAR + cv2.WARP_INVERSE_MAP)
            im5 = cv2.warpPerspective(im5, H, size1, flags=cv2.INTER_LINEAR + cv2.WARP_INVERSE_MAP)
        im1 = cv2.resize(im1, size, interpolation=cv2.INTER_LINEAR)
        im2 = cv2.resize(im2, size, interpolation=cv2.INTER_LINEAR)
        im3 = cv2.resize(im3, size, interpolation=cv2.INTER_LINEAR)
        im4 = cv2.resize(im4, size, interpolation=cv2.INTER_LINEAR)
        im5 = cv2.resize(im5, size, interpolation=cv2.INTER_LINEAR)
        return im1, im2, im3, im4, im5

    def ORB(self, pathA, pathB, pathC):
        img1 = io.imread(pathA)
        img1 = self.RGB2GBR(img1)
        img2 = io.imread(pathB)
        img2 = self.RGB2GBR(img2)
        img3 = io.imread(pathC)
        img3 = self.RGB2GBR(img3)

        size = (img1.shape[1], img2.shape[0])
        im1 = cv2.resize(img1, (1000, 1000), interpolation=cv2.INTER_LINEAR)
        im2 = cv2.resize(img2, (1000, 1000), interpolation=cv2.INTER_LINEAR)
        im3 = cv2.resize(img3, (1000, 1000), interpolation=cv2.INTER_LINEAR)

        result1 = self.siftImageAlignment(im1, im2)
        result2 = self.siftImageAlignment(im1, im3)
        result1 = cv2.resize(result1, size, interpolation=cv2.INTER_LINEAR)
        result2 = cv2.resize(result2, size, interpolation=cv2.INTER_LINEAR)
        image = self.combineRGB(img1, result1, result2)
        return image

    def ORB2(self, pathA, pathB):
        img1 = io.imread(pathA)
        img1 = self.RGB2GBR(img1)
        img2 = io.imread(pathB)
        img2 = self.RGB2GBR(img2)
        size = (img1.shape[1], img2.shape[0])
        im1 = cv2.resize(img1, (1000, 1000), interpolation=cv2.INTER_LINEAR)
        im2 = cv2.resize(img2, (1000, 1000), interpolation=cv2.INTER_LINEAR)
        result = self.siftImageAlignment(im1, im2)
        result = cv2.resize(result, size, interpolation=cv2.INTER_LINEAR)
        return result

    def autoORB(self):
        paths, names = self.get_images_paths_new()
        prePath = osp.dirname(paths[1])
        img0 = io.imread(paths[1])
        size = (img0.shape[1], img0.shape[0])
        size1 = (1000, 1000)
        img0 = cv2.resize(img0, size1, interpolation=cv2.INTER_LINEAR)
        for i in range(len(paths)):
            print('开始对齐：' + paths[i])
            img = io.imread(paths[i])
            # img = self.RGB2GBR(img)
            img = cv2.resize(img, size1, interpolation=cv2.INTER_LINEAR)
            result = self.siftImageAlignment(img0, img)
            img = cv2.resize(result, size, interpolation=cv2.INTER_LINEAR)
            savePath = prePath + '/' + names[i] + 'DQ.jpg'
            # cv2.imwrite(savePath, img)
            io.imsave(savePath, img)
            print('对齐完成：' + names[i])

    # 中文路径读取
    def RGB2GBR(self, image):
        return image[:, :, ::-1]

    # 提取RGB，融合图像
    def combineRGB(self, img1, img2, img3):
        img = cv2.addWeighted(img1, 0.5, img2, 0.5, 0)
        img = cv2.addWeighted(img, 0.5, img3, 0.5, 0)
        return img
        # return np.maximum(np.maximum(img1, img2), img3)

    # 偏光图像对比功能
    def Pimage_contrast(self):
        images, names = self.get_images_pixmap()
        self.imagesList.loadImages(images, names)

    def get_images_paths(self):
        path = osp.dirname(self.imagePath)
        basename = osp.basename(self.imagePath)
        nf = osp.abspath('./icons/404.jpg')
        prefix = '-'.join(basename.split('-')[:2])
        paths = [
            osp.join(path, prefix + '-A-1.jpg'), osp.join(path, prefix + '-B-1.jpg'),
            osp.join(path, prefix + '-A-2.jpg'), osp.join(path, prefix + '-B-2.jpg'),
            osp.join(path, prefix + '-A-3.jpg'), osp.join(path, prefix + '-B-3.jpg'),
            osp.join(path, prefix + '-A-4.jpg'), osp.join(path, prefix + '-B-4.jpg'),
            osp.join(path, prefix + '-A-5.jpg'), osp.join(path, prefix + '-B-5.jpg'),
        ]
        names = [
            prefix + '-A-1', prefix + '-B-1',
            prefix + '-A-2', prefix + '-B-2',
            prefix + '-A-3', prefix + '-B-3',
            prefix + '-A-4', prefix + '-B-4',
            prefix + '-A-5', prefix + '-B-5',
        ]
        paths = list(map(lambda x: x if osp.exists(x) else nf, paths))
        return paths, names

    def get_server_image_paths(self):
        path = './temp/contrast'
        if osp.exists(path):
            shutil.rmtree(path)
        os.makedirs(path)

        # 缓存图像
        prefix = '_'.join(self.filename.split('_')[:5])

        names = [
            prefix + '_A_D_0.JPG',   prefix + '_B_D_0.JPG',
            prefix + '_A_D_36.JPG',  prefix + '_B_D_36.JPG',
            prefix + '_A_D_72.JPG',  prefix + '_B_D_72.JPG',
            prefix + '_A_D_108.JPG', prefix + '_B_D_108.JPG',
            prefix + '_A_D_144.JPG', prefix + '_B_D_144.JPG',
            prefix + '_B_C_36.JPG',
        ]

        paths = [osp.abspath(osp.join(path, name)) for name in names]
        IS_EXISTS_URL = self.server_path + '/DataManagementSystem/isExit'
        DOWNLOAD_URL = self.server_path + '/DataManagementSystem/download'
        for i, name in enumerate(names):
            res = requests.post(
                IS_EXISTS_URL,
                data={
                    'fileName': name,
                    'typeId': 1
                }
            )
            if res.json():
                res = requests.get(
                    DOWNLOAD_URL,
                    params={
                        'fileName': name,
                        'typeId': 1
                    },
                )
                with open(paths[i], 'wb') as f:
                    f.write(res.content)
        nf = osp.abspath('./icons/404.jpg')
        paths = list(map(lambda x: x if osp.exists(x) else nf, paths))
        return paths, names

    def get_images_pixmap(self):
        pixmap = []
        if self.IsServer:
            paths, names = self.get_server_image_paths()
        else:
            paths, names = self.get_images_paths_new()
        for path in paths:
            pixmap.append(QtGui.QPixmap(path))
        return pixmap, names

    def get_images_paths_new(self):
        path = osp.dirname(self.imagePath)
        basename = osp.basename(self.imagePath)
        nf = osp.abspath('./icons/404.jpg')
        prefix = '_'.join(basename.split('_')[:5])

        path_all = glob.glob(osp.join(path, prefix + '*.JPG'))
        path_all = sorted(path_all, key=lambda i: int(i.split('_')[-1].split('.')[0].split('-')[0]))
        paths = []
        for a in path_all:
            paths.append(a)

        names = list(map(lambda x: osp.splitext(osp.basename(x))[0], paths))
        paths = list(map(lambda x: x if osp.exists(x) else nf, paths))
        return paths, names

    # 预处理
    def imgarr_to_qimage(self, img):
        img_data = utils.img_to_data(img)
        qimage = QtGui.QImage.fromData(img_data)
        return qimage

    def img2pixmap(self, img):
        img = self.RGB2GBR(img)
        img_data = utils.img_to_data(img)
        qimage = QtGui.QImage.fromData(img_data)
        pixiamge = QtGui.QPixmap.fromImage(qimage)
        return pixiamge

    # 光照均衡化功能
    def Light_balance(self):
        test = io.imread(self.imagePath)
        test = self.RGB2GBR(test)
        B, G, R = cv2.split(test)  # get single 8-bits channel
        EB = cv2.equalizeHist(B)
        EG = cv2.equalizeHist(G)
        ER = cv2.equalizeHist(R)
        equal_test = cv2.merge((EB, EG, ER))  # merge it back
        img = cv2.cvtColor(equal_test, cv2.COLOR_BGR2RGB)
        # img = self.RGB2GBR(img)
        img_data = utils.img_to_data(img)
        qimage = QtGui.QImage.fromData(img_data)
        self.onNewBrightnessContrast(qimage)
        if self.IsServer:
            # 保存结果图像
            img_path = 'icons\pre-treatment.jpg'
            cv2.imwrite(img_path, utils.img_b64_to_arr(img_data))
            # 上传服务器
            self.upload(img_path, 2)
            # os.remove('ResultsImage/equal_test.jpg')

    # 图像模糊
    def blur(self):
        # if self.preImage is not None:
        #     img = self.preImage
        # else:
        img = Image.open(self.imagePath)
        # 图像模糊
        img = img.filter(ImageFilter.BLUR)
        # self.preImage = img
        img_data = utils.img_pil_to_data(img)
        qimage = QtGui.QImage.fromData(img_data)
        self.onNewBrightnessContrast(qimage)

    # 细节增强滤波
    def detail(self):
        img = Image.open(self.imagePath)
        # 细节增强滤波，会使得图像中细节更加明显。
        img = img.filter(ImageFilter.DETAIL)
        img_data = utils.img_pil_to_data(img)
        qimage = QtGui.QImage.fromData(img_data)
        self.onNewBrightnessContrast(qimage)

    # 边缘增强滤波
    def edge_enhance(self):
        img = Image.open(self.imagePath)
        # 边缘增强滤波，突出、加强和改善图像中不同灰度区域之间的边界和轮廓的图像增强方法。
        # 经处理使得边界和边缘在图像上表现为图像灰度的突变,用以提高人眼识别能力。
        img = img.filter(ImageFilter.EDGE_ENHANCE)
        img_data = utils.img_pil_to_data(img)
        qimage = QtGui.QImage.fromData(img_data)
        self.onNewBrightnessContrast(qimage)

    # 图像锐化
    def sharpen(self):
        img = Image.open(self.imagePath)
        img = img.filter(ImageFilter.SHARPEN)
        # self.preImage = img
        img_data = utils.img_pil_to_data(img)
        qimage = QtGui.QImage.fromData(img_data)
        self.onNewBrightnessContrast(qimage)

    # 高斯滤波
    def gassBlur(self):
        img = Image.open(self.imagePath)
        img = img.filter(ImageFilter.GaussianBlur)
        img_data = utils.img_pil_to_data(img)
        qimage = QtGui.QImage.fromData(img_data)
        self.onNewBrightnessContrast(qimage)

    # 平滑去噪功能
    def Smooth_denoising(self):
        image = Image.open(self.imagePath)
        img = image.filter(ImageFilter.SMOOTH_MORE)
        img_data = utils.img_pil_to_data(img)
        qimage = QtGui.QImage.fromData(img_data)
        self.onNewBrightnessContrast(qimage)
        if self.IsServer:
            # 保存结果图像
            img_path = 'icons/pre-treatment.jpg'
            cv2.imwrite(img_path, utils.img_data_to_arr(img_data))
            # 上传服务器
            self.upload(img_path, 2)
            # os.remove('ResultsImage/bilateral_filter_img1.jpg')

    # 计算颗粒长径
    def lengthDiameter(self, countour):
        rect = cv2.minAreaRect(countour)
        Aspect = rect[1][0]
        if Aspect < rect[1][1]:
            Aspect = rect[1][1]
        return Aspect

    # 传统分割功能
    def auto_segmentation1(self):
        f_name = self.filename
        self.loadFile(self.imagePath)
        if self.IsServer:
            self.filename = f_name
        self.Image_type = 3
        features = self.pelletFeatures

        dir_path = osp.dirname(self.imagePath)
        path_A = glob.glob(dir_path + '/*A_D_*')[0]
        path_B = glob.glob(dir_path + '/*B_D_*')

        '''logging.debug('segment: watershed...')'''
        contours = new_seg_zhimi.new_seg(path_A, path_B)
        print('开始画点:', len(contours))
        for cnts in contours:
            for n, contour in enumerate(cnts):
                if len(contour) > 50:
                    p1 = Thread(target=self.pointDraw, args=[contour])
                    p1.start()
                    p1.join()
                    shape = self.canvas.setLastLabel("石英", self.labelDialog.getFlags())
                    shape.group_id = None
                    # 保存颗粒信息
                    shape.feature_id = n
                    shape.face_area = self.cnt_area(contour)
                    shape.face_porosity = self.lengthDiameter(contour)
                    shape.features = copy.deepcopy(features)
                    self.addLabel(shape)
                    self.setEditMode()

    def auto_segmentation(self):
            f_name = self.filename
            self.loadFile(self.imagePath)
            if self.IsServer:
                self.filename = f_name
            self.Image_type = 3
            features = self.pelletFeatures
            # 分割
            path = osp.dirname(self.imagePath)
            pA = glob.glob(path + '/*A_D_*')[0]
            pBs = glob.glob(path + '/*B_D_*')
            im1 = io.imread(pA)
            im1 = im1[:, :, ::-1]
            im2 = io.imread(pBs[0])
            im2 = self.RGB2GBR(im2)
            im3 = io.imread(pBs[1])
            im3 = self.RGB2GBR(im3)
            im4 = io.imread(pBs[2])
            im4 = self.RGB2GBR(im4)
            im5 = io.imread(pBs[3])
            im5 = self.RGB2GBR(im5)
            im6 = io.imread(pBs[4])
            im6 = self.RGB2GBR(im6)
            ''' logging.debug('segment: watershed....') '''
            contours = segmentation.new_seg(im1, im2, im3, im4, im5, im6)
            print('开始画点:', len(contours))
            for cnts in contours:
                for n, contour in enumerate(cnts):
                    if len(contour) > 50:
                        p1 = Thread(target=self.pointDraw, args=[contour])
                        p1.start()
                        p1.join()
                        shape = self.canvas.setLastLabel("石英", self.labelDialog.getFlags())
                        shape.group_id = None
                        # 保存颗粒信息
                        shape.feature_id = n
                        shape.face_area = self.cnt_area(contour)
                        shape.face_porosity = self.lengthDiameter(contour)
                        shape.features = copy.deepcopy(features)
                        self.addLabel(shape)
                        self.setEditMode()

    def pointDraw(self, contour):
        length = len(contour)
        value = 50 // length
        if value == 0:
            value = length // 3
        step = length // (value * 10)
        if step == 0:
            step = 15
        else:
            step = 30
        for i in range(0, length, step):
            p1 = QtCore.QPointF(contour[i][0][0], contour[i][0][1])
            self.canvas.generatePolygon(p1)
        self.canvas.generatePolygon(QtCore.QPointF(contour[0][0][0], contour[0][0][1]))
        self.canvas.shapes.append(self.canvas.current)
        self.canvas.storeShapes()
        self.canvas.current = None

    # 添加默认标签名字
    def addLabel1(self, shape):
        text = shape.label
        label_list_item = LabelListWidgetItem(text, shape)
        self.labelList.addItem(label_list_item)
        if not self.uniqLabelList.findItemsByLabel(shape.label):
            item = self.uniqLabelList.createItemFromLabel(shape.label)
            self.uniqLabelList.addItem(item)
            rgb = self._get_rgb_by_label(shape.label)
            self.uniqLabelList.setItemLabel(item, shape.label, rgb)
        self.labelDialog.addLabelHistory(shape.label)
        for action in self.actions.onShapesPresent:
            action.setEnabled(True)
        rgb = self._get_rgb_by_label(shape.label)
        r, g, b = rgb
        label_list_item.setText(
            '{} <font color="#{:02x}{:02x}{:02x}">●</font>'.format(
                text, r, g, b
            )
        )
        shape.line_color = QtGui.QColor(r, g, b)
        shape.vertex_fill_color = QtGui.QColor(r, g, b)
        shape.hvertex_fill_color = QtGui.QColor(255, 255, 255)
        shape.fill_color = QtGui.QColor(r, g, b, 128)
        shape.select_line_color = QtGui.QColor(255, 255, 255)
        shape.select_fill_color = QtGui.QColor(r, g, b, 155)

    def auto_recognition(self):
        # print('开始识别：')
        # path = osp.dirname(self.imagePath)
        # pA = glob.glob(path + '/*A_D_0*.JPG')[0]
        # pB1 = glob.glob(path + '/*B_D_0*.JPG')[0]
        # pB2 = glob.glob(path + '/*B_D_36*.JPG')[0]
        # pB3 = glob.glob(path + '/*B_D_72*.JPG')[0]
        # pB4 = glob.glob(path + '/*B_D_108*.JPG')[0]
        # pB5 = glob.glob(path + '/*B_D_144*.JPG')[0]
        # pC = glob.glob(path + '/*_C_*.JPG')[0]
        # json_path = glob.glob(path + '/*.json')[0]
        # im1 = io.imread(pA)
        # im1 = self.RGB2GBR(im1)
        # im2 = io.imread(pB1)
        # im2 = self.RGB2GBR(im2)
        # im3 = io.imread(pB2)
        # im3 = self.RGB2GBR(im3)
        # im4 = io.imread(pB3)
        # im4 = self.RGB2GBR(im4)
        # im5 = io.imread(pB4)
        # im5 = self.RGB2GBR(im5)
        # im6 = io.imread(pB5)
        # im6 = self.RGB2GBR(im6)
        # im7 = io.imread(pC)
        # im7 = self.RGB2GBR(im7)
        # im2, im3, im4, im5, im6 = self.siftImageAlignment1(im1, im2, im3, im4, im5, im6)
        # predict.predict_img(json_path, im1, im2, im3, im4, im5, im6, im7)
        pass

    def auto_recognition1(self):
        features = self.pelletFeatures
        labelsDict = self.labelsDict
        dir = os.path.dirname(self.imagePath)
        r_image = io.imread(dir + '/predict.png')
        shapeId = 1
        for n in range(1, 17):
            r_img = r_image.copy()
            r_img[r_image != n] = 0
            r_img = r_img[:, :, 0]
            dst_img = np.array(r_img, dtype=np.uint8)
            contours = segmentation.contoursCv(dst_img)
            if contours == 0:
                return
            else:
                for i, contour in enumerate(contours):
                    contour = cv2.approxPolyDP(contour, 2, True)
                    p1 = Thread(target=self.pointDraw1, args=[contour])
                    p1.start()
                    p1.join()
                    shape = self.canvas.setLastLabel(labelsDict[n], self.labelDialog.getFlags())
                    shape.group_id = None
                    # 保存颗粒信息p
                    shape.feature_id = shapeId
                    shape.face_area = self.cnt_area(contour)
                    shape.face_porosity = self.lengthDiameter(contour)
                    shapeId = shapeId + 1
                    shape.features = copy.deepcopy(features)
                    self.addLabel(shape)
                    self.setEditMode()

    def pointDraw1(self, contour):
        length = len(contour)  # 单个轮廓的大小
        pos = []
        for i in range(0, length):
            pos = QtCore.QPointF(contour[i][0][0], contour[i][0][1])
            self.canvas.generatePolygon(pos)
        self.canvas.generatePolygon(QtCore.QPointF(contour[0][0][0], contour[0][0][1]))
        self.canvas.shapes.append(self.canvas.current)
        self.canvas.storeShapes()
        self.canvas.current = None

    # 语义分割功能----
    def layer_contrast(self):
        features = self.pelletFeatures
        labelsDict = self.labelsDict
        # deeplab
        fname = os.path.basename(self.imagePath)
        fname = fname.split('.')[0]
        image_files = glob.glob('./Net/temp/' + '*.jpg')
        for imgf in image_files:
            os.remove(imgf)
        resizeImage.Crop(fname, self.imagePath)
        predict.DeepLabPredict.predictDeeplab(self, imagepath='./Net/temp/')
        resizeImage.cat(self.imagePath, './Net/temp/'+fname)

        r_image = cv2.imread('icons/predict.jpg')
        r_image = cv2.cvtColor(r_image, cv2.COLOR_BGR2GRAY)
        contours = segmentation.contoursCv(r_image)
        # contours = contours[0:int(0.6 * len(contours))]
        contours = list(filter(lambda x: len(x) > 100, contours))
        for n, contour in enumerate(contours):
            contour = cv2.approxPolyDP(contour, 2, True)
            p1 = Thread(target=self.pointDraw1, args=[contour])
            p1.start()
            p1.join()
            nClass = r_image[contour[0][0][1], contour[0][0][0]]
            nameLabel = labelsDict[nClass]
            shape = self.canvas.setLastLabel(nameLabel, self.labelDialog.getFlags())
            shape.group_id = None
            # 保存颗粒信息
            shape.feature_id = n
            shape.face_area = self.cnt_area(contour)
            shape.face_porosity = self.lengthDiameter(contour)
            # features['feature_id'] = n  # ID
            # features['facePorosity'] = self.lengthDiameter(contour)  # 长径
            # features['faceArea'] = self.cnt_area(contour)  # 面积
            shape.features = copy.deepcopy(features)
            self.addLabel(shape)
            self.setEditMode()

    # 生成报告功能----
    # 制作表格(孔隙)
    def makeExecelPore(self):
        self.countProps()
        ecxelpath = 'icons/poreTemplate.xls'
        workbook = xlrd.open_workbook(ecxelpath, formatting_info=True)
        w = XLWTWriter()
        # process进行xlrd与xlwt间操作，复制一份。
        process(XLRDReader(workbook, 'unknown.xls'), w)
        # w.output[0][1]为copy后返回的对象
        wb = w.output[0][1]
        # style_list为原文件的单元格格式信息，列表对象
        style_list = w.style_list
        for n, sheet in enumerate(workbook.sheets()):
            sheet2 = wb.get_sheet(n)
            for r in range(sheet.nrows):
                for c, cell in enumerate(sheet.row_values(r)):
                    # 若循环到第一个sheet的（3,6）单元格，修改值
                    if n == 0 and r == 3 and c == 8:
                        values = self.facePorosity()
                    else:
                        values = sheet.cell_value(r, c)
                    style = style_list[sheet.cell_xf_index(r, c)]
                    sheet2.write(r, c, values, style)
        fname = os.path.basename(self.imagePath)
        fname = fname.split('.')[0]
        dir, ftype = QtWidgets.QFileDialog.getSaveFileName(self, "文件保存", fname + '面孔率报告', 'xls(*.xls)')
        if dir == "":
            return
        else:
            wb.save(dir)

    # 制作颗粒粒径表格
    def makeExecelParticl(self):
        self.countProps()
        file_path = 'icons/particl.xls'
        fileobj = xlrd.open_workbook(file_path, formatting_info=True)
        # 准备写入一个新的表
        w = XLWTWriter()
        # process进行xlrd与xlwt间操作，复制一份。
        process(XLRDReader(fileobj, 'unknown.xls'), w)
        # w.output[0][1]为copy后返回的对象
        wb = w.output[0][1]
        # style_list为原文件的单元格格式信息，列表对象
        style_list = w.style_list
        # 获取第一个工作表
        sheet = fileobj.sheets()[0]
        sheet1 = wb.get_sheet(0)
        row_count = sheet.nrows  # 行数
        col_count = sheet.ncols  # 列数
        # 复制一张
        for i in range(row_count):
            for j in range(col_count):
                values = sheet.cell_value(i, j)
                style = style_list[sheet.cell_xf_index(i, j)]
                sheet1.write(i, j, values, style)
        #  定位个数和面积比率位置
        #  定位需要填入的位置
        kl_x = 0
        kl_y = 0
        for element in range(row_count):
            for num in range(col_count):
                if '颗粒个数' in str(sheet.row_values(element)[num]).replace(' ', ''):
                    kl_x = element
                    kl_y = num
        # 传入颗粒属性
        length, area, _ = self.countParticl()
        # 设置插入的格式
        style = style_list[sheet.cell_xf_index(kl_x, kl_y)]
        al = xlwt.Alignment()
        al.horz = 0x02  # 设置水平居中
        al.vert = 0x01  # 设置垂直居中
        style.alignment = al
        # 插入
        for n in range(0, 6):
            sheet1.write(kl_x + 2 + n, kl_y, length[n], style)
            sheet1.write(kl_x + 2 + n, kl_y + 1, area[n], style)
        fname = os.path.basename(self.imagePath)
        fname = fname.split('.')[0]
        dir, ftype = QtWidgets.QFileDialog.getSaveFileName(self, "文件保存", fname+'颗粒粒径报告', 'xls(*.xls)')
        if dir == "":
            return
        else:
            wb.save(dir)

    # 生产定名报告
    def Generate_report2(self):
        self.countProps()
        file_path = 'icons/final.xls'
        fileobj = xlrd.open_workbook(file_path, formatting_info=True)
        # 准备写入一个新的表
        w = XLWTWriter()
        # process进行xlrd与xlwt间操作，复制一份。
        process(XLRDReader(fileobj, 'unknown.xls'), w)
        # w.output[0][1]为copy后返回的对象
        wb = w.output[0][1]
        # style_list为原文件的单元格格式信息，列表对象
        style_list = w.style_list
        # 获取第一个工作表
        sheet = fileobj.sheets()[0]
        sheet1 = wb.get_sheet(0)
        row_count = sheet.nrows  # 行数
        col_count = sheet.ncols  # 列数

        # 复制一张
        for i in range(row_count):
            for j in range(col_count):
                values = sheet.cell_value(i, j)
                style = style_list[sheet.cell_xf_index(i, j)]
                sheet1.write(i, j, values, style)
        dic = self.countFeatures2()
        for key, value in dic.items():  # 在字典中获取需要添加的字符串
            flag = 0
            for element in range(row_count):
                if flag == 1:
                    break
                for num in range(col_count):
                    if key in str(sheet.row_values(element)[num]).replace(' ', ''):
                        if value == '':  # 如果为空则补充
                            style = style_list[sheet.cell_xf_index(element, num)]
                            al = xlwt.Alignment()
                            al.horz = 0x02  # 设置水平居中
                            al.vert = 0x01  # 设置垂直居中
                            style.alignment = al
                            if key in '火山碎屑岩岩块':
                                sheet1.write(element, num + 2, '/', style)
                            elif key in '岩石薄片成分组成分析':
                                sheet1.write(element + 1, num, '/', style)
                            else:
                                sheet1.write(element, num + 1, '/', style)
                        else:
                            style = style_list[sheet.cell_xf_index(element, num)]
                            al = xlwt.Alignment()
                            al.horz = 0x02  # 设置水平居中
                            al.vert = 0x01  # 设置垂直居中
                            style.alignment = al
                            if key in '火山碎屑岩岩块':
                                sheet1.write(element, num + 2, value, style)
                            elif key in '岩石薄片成分组成分析':
                                al.horz = 0x01
                                al.vert = 0x00
                                al.wrap = True
                                sheet1.write(element + 1, num, value, style)
                            else:
                                sheet1.write(element, num + 1, value, style)
                            flag = 1
                            break
        fname = os.path.basename(self.imagePath)
        fname = fname.split('.')[0]
        dir, ftype = QtWidgets.QFileDialog.getSaveFileName(self, "文件保存", fname + '定名报告', 'xls(*.xls)')
        if dir == "":
            return
        else:
            wb.save(dir)

    # 多个多边形融合----
    # 计算欧氏距离
    def Dist(self, p1, p2):
        return np.sqrt((p1[0] - p2[0])**2 + (p1[1] - p2[1])**2)

    def __oDist(self, p1, p2):
        return np.sqrt((p1.x() - p2.x()) ** 2 + (p1.y() - p2.y()) ** 2)

    def merge(self):
        index1, index2 = -1, -1
        dist = -1
        shapes = []
        if len(self.canvas.selectedShapes) >= 2:
            for n in range(len(self.canvas.selectedShapes) - 1):
                shape1: Shape = self.canvas.selectedShapes[0]
                shape2: Shape = self.canvas.selectedShapes[n + 1]
                for i in range(len(shape1.points)):
                    for j in range(len(shape2.points)):
                        d = self.__oDist(shape1.points[i], shape2.points[j])
                        if dist == -1 or d < dist:
                            dist = d
                            index1, index2 = i, j
                k = 0
                for i in range(index2, len(shape2)):
                    shape1.points.insert(index1 + k, shape2[i])
                    k += 1
                for i in range(0, index2):
                    shape1.points.insert(index1 + k, shape2[i])
                    k += 1
                self.canvas.selectedShapes[0] = shape1
                shapes.append(shape2)
        self.canvas.selectedShapes = shapes
        self.remLabels(self.canvas.deleteSelected())
        self.setDirty()
        if self.noShapes():
            for action in self.actions.onShapesPresent:
                action.setEnabled(False)

    # 计算颗粒面积
    def cnt_area(self, cnt):
        area = cv2.contourArea(cnt)
        return area

    # 计算面孔率
    def facePorosity(self):
        img = io.imread(self.imagePath)
        img = self.RGB2GBR(img)
        hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
        low_hsv = np.array([98, 43, 46])
        high_hsv = np.array([124, 255, 255])
        mask = cv2.inRange(hsv, lowerb=low_hsv, upperb=high_hsv)
        # 阈值处理，变成只有黑白的图，返回ret,thersh利用返回的thersh
        # binary = cv2.threshold(mask, 127, 255, cv2.THRESH_BINARY)[1]
        # a = Counter(binary.flatten())[255]
        # b = a + Counter(binary.flatten())[0]
        # x2 = a / b
        # frate = '%.2f%%' % (x2 * 100)
        # return frate
        x = Counter(mask.flatten())
        a = x[0]
        b = x[255]
        x2 = b/(a+b)
        frate = '%.2f%%' % (x2 * 100)
        return frate

    # 统计标签名
    def countLabel(self):
        shapes = self.canvas.shapes
        # 统计各类别数量
        sName = [shape.label for shape in shapes]
        # 用字典存储：字符：次数
        result = {}
        for word in sName:
            if word in result:
                result[word] = result[word] + 1  # 该字符第N次在字典里
            else:
                result[word] = 1  # 该字符第一次在字典里
        # 按照次数从大到小排序
        fin_result = list(zip(result.values(), result.keys()))
        fin_result.sort(reverse=True, key=lambda x: x[0])
        return fin_result

    # 统计每个颗粒的长径和面积
    def countProps(self):
        shapes = self.canvas.shapes
        for shape in shapes:
            count = [[[round(point.x()), round(point.y())]] for point in shape.points]
            count = np.array(count)
            shape.face_porosity = self.lengthDiameter(count)
            shape.face_area = self.cnt_area(count)

    # 统计颗粒的粒径
    def countParticl(self):
        self.countProps()
        length = [0, 0, 0, 0, 0, 0]
        area = [0, 0, 0, 0, 0, 0]
        # basename = osp.basename(self.imagePath)
        # x = basename.split('_')[4]
        # x = x[:-1]
        x = 50
        xscale = 0
        if x == 50:
            xscale = 0.51
        if x == 100:
            xscale = 1
        if x == 150:
            xscale = 1.5
        listA = [shape.face_area for shape in self.canvas.shapes]
        listP = [shape.face_porosity * xscale for shape in self.canvas.shapes]
        # listA = [shape.features['faceArea'] for shape in self.canvas.shapes]
        # listP = [shape.features['facePorosity'] * .00051 for shape in self.canvas.shapes]
        # 对传入的数据进行处理
        for index, k in enumerate(listP):
            if (k > 2.0) and k <= 256.0:
                length[0] += 1
                area[0] += listA[index]
            elif (k > 0.5) and k <= 2:
                length[1] += 1
                area[1] += listA[index]
            elif (k > 0.25) and k <= 0.5:
                length[2] += 1
                area[2] += listA[index]
            elif (k > 0.125) and k <= 0.25:
                length[3] += 1
                area[3] += listA[index]
            elif (k > 0.0625) and k <= 0.125:
                length[4] += 1
                area[4] += listA[index]
            elif (k >= 0.156) and k <= 0.625:
                length[5] += 1
                area[5] += listA[index]
        # 求比例
        img = io.imread(self.imagePath)
        all_area = img.shape[0] * img.shape[1]
        for m in range(0, 6):
            area[m] /= all_area
            area[m] = area[m] * 100
            area[m] = '%.2f' % area[m]
        print(length, area, listP)
        return length, area, listP


    def insert_img(self, sheet, sheet2, dic, k):
        if k == 0:
            img = image1('icons/pic1.png')
            img.width = 110
            img.height = 110
            sheet2.add_image(img, 'F3')
        elif k == 1:
            img1 = image1('icons/check1.png')
            img1.width = 260
            img1.height = 40
            sheet2.add_image(img1, 'D1')
        elif k == 2:
            # 填充数据，插入图片
            sheet2 = self.fill_excl(sheet, sheet2, dic)

            imgth, _ = self.get_images_paths_new()
            img = image1(imgth[0])
            img.width = 358
            img.height = 228
            img1 = image1(imgth[1])
            img1.width = 358
            img1.height = 228

            # img2 = image1('icons\p3.png')
            # img2.width = 358
            # img2.height = 228
            # img3 = image1('icons\p4.png')
            # img3.width = 358
            # img3.height = 228

            img4 = image1('icons/check1.png')
            img4.width = 300
            img4.height = 50
            img5 = image1('icons/c_name1.png')
            img5.width = 200
            img5.height = 70

            sheet2.add_image(img, 'AC7')
            sheet2.cell(12, 29, dic['p1'])
            sheet2.add_image(img1, 'AK7')
            sheet2.cell(12, 37, dic['p2'])
            # sheet2.cell(18, 29, dic['p3'])
            # sheet2.cell(18, 37, dic['p4'])
            # sheet2.add_image(img2, 'AC14')
            # sheet2.add_image(img3, 'AK14')

            sheet2.add_image(img5, 'A1')
            sheet2.add_image(img4, 'H1')
